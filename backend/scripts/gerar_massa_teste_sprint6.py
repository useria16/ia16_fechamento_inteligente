"""
Gerador de massa de teste para Sprint 6 — Motor de conciliação.

Cria 3 planilhas sintéticas preservando a estrutura dos arquivos originais da Daxx,
mas com dados completamente fictícios e marcadores de cenário.

Uso:
  cd backend
  python -m scripts.gerar_massa_teste_sprint6

Saída em: ../docs/massa-teste/
"""
import os
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SAIDA = Path(__file__).parent.parent.parent / "docs" / "massa-teste"
SAIDA.mkdir(parents=True, exist_ok=True)

# Paleta de cores dos cenários para rastreamento visual
COR_MATCH        = "C6EFCE"  # verde claro
COR_DIV_DATA     = "FFEB9C"  # amarelo
COR_PREVISTO_NR  = "BDD7EE"  # azul claro
COR_REALIZADO_NP = "FCE4D6"  # laranja claro
COR_DIV_VALOR    = "F4CCCC"  # vermelho claro
COR_ANTECIPACAO  = "D9EAD3"  # verde médio
COR_ATRASO       = "FFF2CC"  # amarelo escuro
COR_DUP_EXTRATO  = "EAD1DC"  # rosa
COR_DUP_FLUXO    = "D0E0E3"  # ciano claro
COR_PENDENTE     = "CFE2F3"  # azul médio
COR_CABECALHO    = "1F497D"  # azul escuro
COR_TOTAL        = "D9D9D9"  # cinza

def fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def bold_font(size=10, color="000000", name="Calibri"):
    return Font(bold=True, size=size, color=color, name=name)

def header_font():
    return Font(bold=True, size=10, color="FFFFFF", name="Calibri")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ═══════════════════════════════════════════════════════════════════════════════
# Dados dos cenários
# ═══════════════════════════════════════════════════════════════════════════════

# Período de teste: 10/06/2026 a 14/06/2026
D10 = date(2026, 6, 10)
D11 = date(2026, 6, 11)
D12 = date(2026, 6, 12)
D13 = date(2026, 6, 13)
D14 = date(2026, 6, 14)
D15 = date(2026, 6, 15)  # usado no fluxo (cenário 6)
D09 = date(2026, 6, 9)   # saldo anterior

# Lançamentos do extrato bancário
# (data, tipo_lancamento, razao_social, cpf_cnpj, valor, cenario_tag, cor)
EXTRATO_LANCAMENTOS = [
    # ── CENÁRIO 1a: Match perfeito — pagamento ───────────────────────────────
    (D10, "PIX ENVIADO",
     "CENARIO_01_MATCH_PERFEITO SERVICOS LTDA",
     "11.111.111/0001-11", -1500.00, "CEN01a", COR_MATCH),
    # ── CENÁRIO 1b: Match perfeito — recebimento ─────────────────────────────
    (D10, "RECEBIMENTOS",
     "CENARIO_01B_MATCH_PERFEITO COMERCIO SA",
     "22.222.222/0001-22", 8200.00, "CEN01b", COR_MATCH),
    # ── CENÁRIO 2: Valor igual, data diferente (extrato = 11/06, fluxo = 10/06)
    (D11, "BOLETO PAGO",
     "CENARIO_02_DATA_DIFERENTE TECNOLOGIA LTDA",
     "33.333.333/0001-33", -320.00, "CEN02", COR_DIV_DATA),
    # ── CENÁRIO 4: Realizado não previsto ────────────────────────────────────
    (D11, "PIX ENVIADO",
     "CENARIO_04_REALIZADO_NAO_PREVISTO AUTONOMO",
     "444.444.444-44", -950.00, "CEN04", COR_REALIZADO_NP),
    # ── CENÁRIO 5: Valor diferente (extrato=-3400, fluxo=-3600) ──────────────
    (D12, "PIX ENVIADO",
     "CENARIO_05_VALOR_DIFERENTE RECURSOS LTDA",
     "55.555.555/0001-55", -3400.00, "CEN05", COR_DIV_VALOR),
    # ── CENÁRIO 6: Antecipação de recebimento (extrato=12/06, fluxo=15/06) ───
    (D12, "RECEBIMENTOS",
     "CENARIO_06_ANTECIPACAO_RECEBIMENTO CLIENTE SA",
     "66.666.666/0001-66", 5500.00, "CEN06", COR_ANTECIPACAO),
    # ── CENÁRIO 7: Pagamento atrasado (extrato=13/06, fluxo=10/06) ───────────
    (D13, "PIX ENVIADO",
     "CENARIO_07_PAGAMENTO_ATRASADO FORNECEDOR ME",
     "77.777.777/0001-77", -780.00, "CEN07", COR_ATRASO),
    # ── CENÁRIO 8a: Duplicidade extrato (mesmo lançamento 2x) ────────────────
    (D13, "PIX ENVIADO",
     "CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME",
     "888.888.888-88", -200.00, "CEN08a", COR_DUP_EXTRATO),
    # ── CENÁRIO 8b: Duplicidade extrato (cópia) ──────────────────────────────
    (D13, "PIX ENVIADO",
     "CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME",
     "888.888.888-88", -200.00, "CEN08b", COR_DUP_EXTRATO),
    # ── CENÁRIO 9: Duplicidade no fluxo (extrato tem 1, fluxo tem 2) ─────────
    (D14, "BOLETO PAGO",
     "CENARIO_09_DUPLICIDADE_FLUXO SOFTWARE LTDA",
     "99.999.999/0001-99", -650.00, "CEN09", COR_DUP_FLUXO),
    # ── CENÁRIO 10: Pendente análise manual ──────────────────────────────────
    (D14, "TED RECEBIDA",
     "CENARIO_10_PENDENTE_ANALISE TRANSF S/IDENTIFICACAO",
     "00.000.000/0000-00", 3700.00, "CEN10", COR_PENDENTE),
    # ── CENÁRIO 1c: Match perfeito adicional ─────────────────────────────────
    (D14, "RECEBIMENTOS",
     "CENARIO_01C_MATCH_PERFEITO ADICIONAL COMERCIO",
     "10.101.010/0001-01", 2100.00, "CEN01c", COR_MATCH),
]

# Saldos diários calculados
SALDO_INICIAL = 25000.00
saldos = {D09: SALDO_INICIAL}
saldo_atual = SALDO_INICIAL
dias = [D10, D11, D12, D13, D14]
movimentos_dia = {d: [] for d in dias}
for lancamento in EXTRATO_LANCAMENTOS:
    movimentos_dia[lancamento[0]].append(lancamento[4])
for dia in dias:
    for v in movimentos_dia[dia]:
        saldo_atual += v
    saldos[dia] = round(saldo_atual, 2)

# ═══════════════════════════════════════════════════════════════════════════════
# 1. TESTE_Extrato_Lancamentos_banco.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def criar_extrato():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # Metadados do banco (preservar posição exata do original)
    meta = [
        (2, "Atualização:", "10/06/2026 08:00:00"),
        (3, "Nome:",         "EMPRESA SINTESE LTDA"),
        (4, "Agência:",      "0001"),
        (5, "Conta:",        "0099999-1"),
    ]
    for row, label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=row, column=2, value=value)

    ws.cell(row=7, column=1, value="Lançamentos").font = Font(bold=True, name="Calibri", size=12)
    ws.cell(row=8, column=1, value="Periodo:")
    ws.cell(row=8, column=2, value="09/06/2026 até 14/06/2026")

    # Cabeçalho na linha 10 (igual ao original)
    cabecalhos = ["Data", "Lançamento", "Razão Social", "CPF/CNPJ", "Valor (R$)", "Saldo (R$)"]
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=10, column=col, value=cab)
        cell.fill = fill(COR_CABECALHO)
        cell.font = header_font()
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")

    larguras = [12, 30, 45, 22, 14, 14]
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    linha = 11

    # Saldo anterior
    for col_idx, val in enumerate([D09.strftime("%d/%m/%Y"), "SALDO ANTERIOR", "", "", "", saldos[D09]], 1):
        cell = ws.cell(row=linha, column=col_idx, value=val)
        cell.font = Font(bold=True, italic=True, name="Calibri", size=10)
        cell.fill = fill(COR_TOTAL)
    linha += 1

    # Lançamentos por dia
    dia_atual = None
    for lancamento in EXTRATO_LANCAMENTOS:
        data_lanc, tipo, razao, cpf_cnpj, valor, cenario, cor = lancamento

        # Inserir saldo do dia anterior ao mudar de dia
        if dia_atual and dia_atual != data_lanc:
            for col_idx, val in enumerate([
                dia_atual.strftime("%d/%m/%Y"), "SALDO TOTAL DISPONÍVEL DIA",
                "", "", "", saldos[dia_atual]
            ], 1):
                cell = ws.cell(row=linha, column=col_idx, value=val)
                cell.font = Font(bold=True, italic=True, name="Calibri", size=10)
                cell.fill = fill(COR_TOTAL)
            linha += 1

        dia_atual = data_lanc

        # Lançamento
        valores_row = [data_lanc.strftime("%d/%m/%Y"), tipo, razao, cpf_cnpj, valor, ""]
        for col_idx, val in enumerate(valores_row, 1):
            cell = ws.cell(row=linha, column=col_idx, value=val)
            cell.fill = fill(cor)
            cell.border = thin_border()
            if col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            # Adiciona cenário na coluna G (extra para rastreamento)
        ws.cell(row=linha, column=7, value=cenario).font = Font(italic=True, color="808080", size=9)
        linha += 1

    # Saldo do último dia
    if dia_atual:
        for col_idx, val in enumerate([
            dia_atual.strftime("%d/%m/%Y"), "SALDO TOTAL DISPONÍVEL DIA",
            "", "", "", saldos[dia_atual]
        ], 1):
            cell = ws.cell(row=linha, column=col_idx, value=val)
            cell.font = Font(bold=True, italic=True, name="Calibri", size=10)
            cell.fill = fill(COR_TOTAL)
        linha += 1

    ws.column_dimensions["G"].width = 20

    caminho = SAIDA / "TESTE_Extrato_Lancamentos_banco.xlsx"
    wb.save(caminho)
    print(f"  ✔ {caminho.name} ({caminho.stat().st_size // 1024} KB)")
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TESTE_Fluxo_Caixa_Daxx.xlsx
# ═══════════════════════════════════════════════════════════════════════════════

# Categorias do fluxo com seus valores por dia
# Estrutura: (nivel, nome_categoria, {date: valor}, cor_cenario)
FLUXO_CATEGORIAS = [
    # RECEITAS
    (1,  "RECEITAS",                                          {}, None),
    ("", "Outras Receitas",                                   {}, None),
    ("", "Receita Operacional - VENDAS (CLIENTE) - CENARIO_01B_MATCH_PERFEITO",
         {D10: 8200.00},  COR_MATCH),
    ("", "Receita Operacional - VENDAS (CLIENTE) - CENARIO_06_ANTECIPACAO",
         {D15: 5500.00},  COR_ANTECIPACAO),
    ("", "Receita Operacional - VENDAS (CLIENTE) - CENARIO_01C_MATCH_PERFEITO",
         {D14: 2100.00},  COR_MATCH),
    ("", "TOTAL RECEITAS",                                    {
        D10: 8200.00, D14: 2100.00, D15: 5500.00
    }, None),
    # DESPESAS COM PESSOAL
    (1,  "DESPESAS COM PESSOAL",                              {}, None),
    ("", "Folha de Pagamento (PJ) - CENARIO_01_MATCH_PERFEITO",
         {D10: -1500.00}, COR_MATCH),
    ("", "Folha de Pagamento (PJ) - CENARIO_05_VALOR_DIFERENTE",
         {D12: -3600.00}, COR_DIV_VALOR),
    ("", "TOTAL DESPESA COM PESSOAL",                         {
        D10: -1500.00, D12: -3600.00
    }, None),
    # DESPESAS COM FREELANCER
    (1,  "DESPESAS COM FREELANCER",                           {}, None),
    ("", "Freelancer Variável - CENARIO_03_PREVISTO_NAO_REALIZADO",
         {D14: -400.00},  COR_PREVISTO_NR),
    ("", "TOTAL DESPESA FREELANCER",                          {D14: -400.00}, None),
    # DESPESAS DE INFRAESTRUTURA
    (1,  "DESPESAS DE INFRAESTRUTURA",                        {}, None),
    ("", "Internet - CENARIO_02_DATA_DIFERENTE",
         {D10: -320.00},  COR_DIV_DATA),
    ("", "Aluguel PE - CENARIO_07_PAGAMENTO_ATRASADO",
         {D10: -780.00},  COR_ATRASO),
    ("", "Limpeza / Diarista - CENARIO_08_DUPLICIDADE_EXTRATO",
         {D13: -200.00},  COR_DUP_EXTRATO),
    ("", "TOTAL DESPESAS DE INFRAESTRUTURA",                  {
        D10: -1100.00, D13: -200.00
    }, None),
    # DESPESAS DE SOFTWARE
    (1,  "DESPESAS DE SOFTWARE",                              {}, None),
    ("", "Licenças de Software - CENARIO_09A_DUPLICIDADE_FLUXO",
         {D14: -650.00},  COR_DUP_FLUXO),
    ("", "Licenças de Software - CENARIO_09B_DUPLICIDADE_FLUXO",
         {D14: -650.00},  COR_DUP_FLUXO),
    ("", "TOTAL DESPESAS SOFTWARE",                           {D14: -1300.00}, None),
]

DATAS_FLUXO = [D09, D10, D11, D12, D13, D14, D15]


def criar_fluxo_caixa():
    wb = openpyxl.Workbook()

    # ── Aba DAXX MIDIA PE ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DAXX MIDIA PE"

    # Linha 1: título + datas (estrutura transposta do original)
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="EMPRESA SINTESE PE - cód. 9999").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=1, column=3, value="")

    for col_offset, dt in enumerate(DATAS_FLUXO):
        col = 4 + col_offset
        cell = ws.cell(row=1, column=col, value=datetime(dt.year, dt.month, dt.day))
        cell.number_format = "DD/MM/YYYY"
        cell.font = Font(bold=True, name="Calibri", size=9)
        cell.fill = fill(COR_CABECALHO)
        cell.font = Font(bold=True, name="Calibri", size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Coluna ACUMULADO
    col_acumulado = 4 + len(DATAS_FLUXO)
    ws.cell(row=1, column=col_acumulado, value="ACUMULADO PERÍODO").font = Font(bold=True, name="Calibri", size=9)

    # Linha 2 vazia (separador como no original)

    # Categorias a partir da linha 3
    linha = 3
    for nivel, nome, valores, cor in FLUXO_CATEGORIAS:
        is_total = "TOTAL" in nome.upper()
        is_grupo = nivel == 1

        cell_nivel = ws.cell(row=linha, column=1, value=str(nivel) if nivel != "" else "")
        cell_nome  = ws.cell(row=linha, column=2, value=nome)

        if is_grupo:
            cell_nome.font  = Font(bold=True, name="Calibri", size=10)
            cell_nome.fill  = fill("D9D9D9")
            cell_nivel.fill = fill("D9D9D9")
        elif is_total:
            cell_nome.font  = Font(bold=True, name="Calibri", size=10)
            cell_nome.fill  = fill(COR_TOTAL)
        elif cor:
            cell_nome.fill  = fill(cor)
            cell_nome.font  = Font(name="Calibri", size=10)

        acumulado = 0.0
        for col_offset, dt in enumerate(DATAS_FLUXO):
            col  = 4 + col_offset
            val  = valores.get(dt)
            cell = ws.cell(row=linha, column=col, value=val if val is not None else None)
            if val is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                acumulado += val
                if cor:
                    cell.fill = fill(cor)
                if is_total:
                    cell.font = Font(bold=True, name="Calibri", size=10)
                    cell.fill = fill(COR_TOTAL)

        if acumulado != 0:
            cell_ac = ws.cell(row=linha, column=col_acumulado, value=acumulado)
            cell_ac.number_format = '#,##0.00'
            cell_ac.alignment = Alignment(horizontal="right")
            if is_total:
                cell_ac.font = Font(bold=True, name="Calibri", size=10)
                cell_ac.fill = fill(COR_TOTAL)

        ws.cell(row=linha, column=3, value="")  # separador col C
        linha += 1
        if is_grupo:
            linha  # não pular linha extra como no original

    # Larguras
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 3
    for col_offset in range(len(DATAS_FLUXO) + 1):
        ws.column_dimensions[get_column_letter(4 + col_offset)].width = 14

    # ── Aba SALDOS ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="SALDOS")
    ws2.cell(row=2, column=3, value="EMPRESA SINTESE PE").font = Font(bold=True, name="Calibri", size=10)

    cabs_saldo = ["DATA", "", "SALDO FINAL CC", "SALDO FINAL APLICAÇÃO", "TOTAL CC + APLIC"]
    for col, cab in enumerate(cabs_saldo, 1):
        cell = ws2.cell(row=3, column=col, value=cab)
        if cab:
            cell.font  = Font(bold=True, name="Calibri", size=10)
            cell.fill  = fill(COR_CABECALHO)
            cell.font  = Font(bold=True, name="Calibri", size=10, color="FFFFFF")

    saldo_cc = SALDO_INICIAL
    for row_offset, dt in enumerate([D09] + dias):
        movs = sum(movimentos_dia.get(dt, []))
        saldo_cc = round(saldo_cc + movs, 2) if row_offset > 0 else SALDO_INICIAL
        row = 4 + row_offset
        ws2.cell(row=row, column=1, value=datetime(dt.year, dt.month, dt.day)).number_format = "DD/MM/YYYY"
        ws2.cell(row=row, column=3, value=saldo_cc).number_format = "#,##0.00"
        ws2.cell(row=row, column=4, value=0).number_format = "#,##0.00"
        ws2.cell(row=row, column=5, value=saldo_cc).number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 22

    caminho = SAIDA / "TESTE_Fluxo_Caixa_Daxx.xlsx"
    wb.save(caminho)
    print(f"  ✔ {caminho.name} ({caminho.stat().st_size // 1024} KB)")
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# 3. TESTE_Conciliacao_Interna_Referencia.xlsx  (gabarito visual)
# ═══════════════════════════════════════════════════════════════════════════════

# Dados do gabarito visual: o que o time faria manualmente
# (data, desc_banco, desc_fornecedor_cliente, nf_doc, valor_nf, entrada, saida, saldo, resultado_esperado, cor)
CONCILIACAO_LINHAS = []

saldo_running = SALDO_INICIAL
# Saldo anterior
CONCILIACAO_LINHAS.append((
    D09, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, saldo_running,
    "saldo_inicial", COR_TOTAL
))

# Dia 10/06
saldo_running += -1500.00
CONCILIACAO_LINHAS.append((D10, "PIX ENVIADO", "CENARIO_01_MATCH_PERFEITO SERVICOS LTDA - Folha PJ Mai/26", "001", 1500.00, 0, 1500.00, round(saldo_running, 2), "conciliado", COR_MATCH))
saldo_running += 8200.00
CONCILIACAO_LINHAS.append((D10, "RECEBIMENTOS", "CENARIO_01B_MATCH_PERFEITO COMERCIO SA - NF 2026-001", "2026-001", 8200.00, 8200.00, 0, round(saldo_running, 2), "conciliado", COR_MATCH))
CONCILIACAO_LINHAS.append((D10, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, round(saldo_running, 2), "saldo_dia", COR_TOTAL))

# Dia 11/06
saldo_running += -320.00
CONCILIACAO_LINHAS.append((D11, "BOLETO PAGO", "CENARIO_02_DATA_DIFERENTE TECNOLOGIA LTDA - Fatura Internet Jun/26", "INV-0320", 320.00, 0, 320.00, round(saldo_running, 2), "divergencia_data (previsto 10/06, realizado 11/06)", COR_DIV_DATA))
saldo_running += -950.00
CONCILIACAO_LINHAS.append((D11, "PIX ENVIADO", "CENARIO_04_REALIZADO_NAO_PREVISTO AUTONOMO - Serv. eventual", "S/NF", 950.00, 0, 950.00, round(saldo_running, 2), "realizado_nao_previsto", COR_REALIZADO_NP))
CONCILIACAO_LINHAS.append((D11, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, round(saldo_running, 2), "saldo_dia", COR_TOTAL))

# Dia 12/06
saldo_running += -3400.00
CONCILIACAO_LINHAS.append((D12, "PIX ENVIADO", "CENARIO_05_VALOR_DIFERENTE RECURSOS LTDA - Folha PJ Jun/26 (previsto 3.600)", "FOLHA-JUN", 3600.00, 0, 3400.00, round(saldo_running, 2), "divergencia_valor (prev=-3600, real=-3400)", COR_DIV_VALOR))
saldo_running += 5500.00
CONCILIACAO_LINHAS.append((D12, "RECEBIMENTOS", "CENARIO_06_ANTECIPACAO_RECEBIMENTO CLIENTE SA - NF 2026-066 (prev 15/06)", "2026-066", 5500.00, 5500.00, 0, round(saldo_running, 2), "divergencia_data / antecipacao (previsto 15/06)", COR_ANTECIPACAO))
CONCILIACAO_LINHAS.append((D12, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, round(saldo_running, 2), "saldo_dia", COR_TOTAL))

# Dia 13/06
saldo_running += -780.00
CONCILIACAO_LINHAS.append((D13, "PIX ENVIADO", "CENARIO_07_PAGAMENTO_ATRASADO FORNECEDOR ME - Aluguel Jun/26 (prev 10/06)", "ALG-JUN", 780.00, 0, 780.00, round(saldo_running, 2), "divergencia_data / atraso (previsto 10/06)", COR_ATRASO))
saldo_running += -200.00
CONCILIACAO_LINHAS.append((D13, "PIX ENVIADO", "CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME - Diarista 1a.quinz", "80", 200.00, 0, 200.00, round(saldo_running, 2), "conciliado (1a. ocorrência)", COR_DUP_EXTRATO))
saldo_running += -200.00
CONCILIACAO_LINHAS.append((D13, "PIX ENVIADO", "CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME - DUPLICADO - verificar", "80-DUP", 200.00, 0, 200.00, round(saldo_running, 2), "duplicidade_extrato (2a. ocorrência sem cobertura no fluxo)", COR_DUP_EXTRATO))
CONCILIACAO_LINHAS.append((D13, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, round(saldo_running, 2), "saldo_dia", COR_TOTAL))

# Dia 14/06
saldo_running += -650.00
CONCILIACAO_LINHAS.append((D14, "BOLETO PAGO", "CENARIO_09_DUPLICIDADE_FLUXO SOFTWARE LTDA - Lic. Software Jun/26", "SOFT-001", 650.00, 0, 650.00, round(saldo_running, 2), "duplicidade_fluxo (extrato tem 1, fluxo tem 2 linhas de -650)", COR_DUP_FLUXO))
saldo_running += 3700.00
CONCILIACAO_LINHAS.append((D14, "TED RECEBIDA", "CENARIO_10_PENDENTE_ANALISE - Transferência não identificada", "S/REF", 3700.00, 3700.00, 0, round(saldo_running, 2), "pendente_analise_manual", COR_PENDENTE))
saldo_running += 2100.00
CONCILIACAO_LINHAS.append((D14, "RECEBIMENTOS", "CENARIO_01C_MATCH_PERFEITO ADICIONAL COMERCIO - NF 2026-010", "2026-010", 2100.00, 2100.00, 0, round(saldo_running, 2), "conciliado", COR_MATCH))
CONCILIACAO_LINHAS.append((D14, "SALDO TOTAL DISPONÍVEL DIA", "", "", "", 0, 0, round(saldo_running, 2), "saldo_dia", COR_TOTAL))

# Previsto não realizado (apenas referência — não aparece no extrato)
CONCILIACAO_LINHAS.append((D14, "[PREVISTO NÃO REALIZADO]",
    "CENARIO_03_PREVISTO_NAO_REALIZADO AUTONOMO - Freelancer Jun/26 (previsto, não pago)",
    "PEND", 400.00, 0, 0, 0, "previsto_nao_realizado", COR_PREVISTO_NR))


def criar_conciliacao_interna():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maio26"  # preservar nome da aba do original

    # Metadados (mesma estrutura do original)
    ws.cell(row=1, column=1, value=datetime(2026, 6, 10)).number_format = "DD/MM/YYYY"
    ws.cell(row=1, column=2, value="REFERENCIA GABARITO MASSA TESTE")

    ws.cell(row=6,  column=1, value="Atualização:").font = Font(bold=True)
    ws.cell(row=6,  column=2, value="10/06/2026 às 08:00:00h")
    ws.cell(row=7,  column=1, value="Nome:").font = Font(bold=True)
    ws.cell(row=7,  column=2, value="EMPRESA SINTESE LTDA")
    ws.cell(row=8,  column=1, value="Agência:").font = Font(bold=True)
    ws.cell(row=8,  column=2, value="001")
    ws.cell(row=9,  column=1, value="Conta:").font = Font(bold=True)
    ws.cell(row=9,  column=2, value="99999")
    ws.cell(row=11, column=1, value="Periodo:  Junho/2026").font = Font(bold=True)

    # Cabeçalho linha 13 (preservar igual ao original)
    cabs = ["DATA", "DESCRIÇÃO LANÇAMENTO BANCO", "DESCRIÇÃO FORNECEDOR/CLIENTE",
            "NF / DOC", "VALOR NF/DOC", "ENTRADA EXTRATO", "SAIDA EXTRATO", "SALDO",
            "RESULTADO ESPERADO"]
    for col, cab in enumerate(cabs, 1):
        cell = ws.cell(row=13, column=col, value=cab)
        cell.fill  = fill(COR_CABECALHO)
        cell.font  = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()

    # Dados
    for row_offset, (data, desc_banco, desc_forn, nf_doc, valor_nf, entrada, saida, saldo, resultado, cor) in enumerate(CONCILIACAO_LINHAS):
        row = 14 + row_offset
        is_saldo = "SALDO" in desc_banco or "[PREVISTO" in desc_banco
        valores_row = [
            datetime(data.year, data.month, data.day),
            desc_banco, desc_forn, nf_doc,
            valor_nf if valor_nf else None,
            entrada if entrada else None,
            saida if saida else None,
            saldo if saldo else None,
            resultado,
        ]
        for col, val in enumerate(valores_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = fill(cor)
            cell.border = thin_border()
            cell.font   = Font(name="Calibri", size=10,
                               bold=(is_saldo or "[PREVISTO" in desc_banco))
            if col == 1 and isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"
            if col in (5, 6, 7, 8) and val:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Larguras
    widths = [12, 30, 48, 14, 14, 16, 16, 14, 42]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    caminho = SAIDA / "TESTE_Conciliacao_Interna_Referencia.xlsx"
    wb.save(caminho)
    print(f"  ✔ {caminho.name} ({caminho.stat().st_size // 1024} KB)")
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# 4. GABARITO_MASSA_TESTE_DAXX.md
# ═══════════════════════════════════════════════════════════════════════════════
GABARITO_MD = """\
# Gabarito da Massa de Teste — Sprint 6

## Visão geral

| Item | Valor |
|---|---|
| Empresa fictícia | EMPRESA SINTESE LTDA |
| Conta bancária | 0099999-1 / Agência 0001 |
| Período do teste | 10/06/2026 a 14/06/2026 |
| Schema de destino | ia16_fechamento_dev |
| Bucket de destino | arquivos-originais |
| Tolerância de data | 3 dias (configurável) |
| Tolerância de valor | 1,00% do valor (configurável) |

---

## Cenários e registros esperados

---

### CENARIO_01 — Match perfeito

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 10/06/2026 | 10/06/2026 |
| Razão Social / Categoria | CENARIO_01_MATCH_PERFEITO SERVICOS LTDA | Folha de Pagamento (PJ) - CENARIO_01_MATCH_PERFEITO |
| Valor | -1.500,00 | -1.500,00 |
| NF/DOC | 001 | — |

**Resultado esperado:** `conciliado`

**Observação:** Data igual, valor igual, categoria compatível. Par perfeito. O motor deve associar automaticamente.

---

### CENARIO_01B — Match perfeito (receita)

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 10/06/2026 | 10/06/2026 |
| Razão Social / Categoria | CENARIO_01B_MATCH_PERFEITO COMERCIO SA | Receita Operacional - VENDAS (CLIENTE) - CENARIO_01B |
| Valor | +8.200,00 | +8.200,00 |

**Resultado esperado:** `conciliado`

**Observação:** Match em receita. Mesmo comportamento do CENARIO_01.

---

### CENARIO_01C — Match perfeito (adicional)

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 14/06/2026 | 14/06/2026 |
| Razão Social / Categoria | CENARIO_01C_MATCH_PERFEITO ADICIONAL COMERCIO | Receita Operacional - VENDAS (CLIENTE) - CENARIO_01C |
| Valor | +2.100,00 | +2.100,00 |

**Resultado esperado:** `conciliado`

---

### CENARIO_02 — Valor igual, data diferente

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | **11/06/2026** | **10/06/2026** |
| Razão Social / Categoria | CENARIO_02_DATA_DIFERENTE TECNOLOGIA LTDA | Internet - CENARIO_02_DATA_DIFERENTE |
| Valor | -320,00 | -320,00 |
| Diferença de data | 1 dia | — |

**Resultado esperado:** `divergencia_data`

**Tolerância:** Dentro de 3 dias → deve ser sinalizado como divergência de data, mas ainda conciliável com confirmação.

**Observação:** O boleto foi pago 1 dia após o previsto. Valor idêntico. O motor deve encontrar a correspondência e marcar a data como divergente.

---

### CENARIO_03 — Previsto não realizado

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | — (não existe) | 14/06/2026 |
| Razão Social / Categoria | — | Freelancer Variável - CENARIO_03_PREVISTO_NAO_REALIZADO |
| Valor | — | -400,00 |

**Resultado esperado:** `previsto_nao_realizado`

**Observação:** Item estava planejado no fluxo para 14/06/2026, mas não gerou lançamento bancário. Motor deve listar como não conciliado no lado do fluxo.

---

### CENARIO_04 — Realizado não previsto

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 11/06/2026 | — (não existe) |
| Razão Social / Categoria | CENARIO_04_REALIZADO_NAO_PREVISTO AUTONOMO | — |
| Valor | -950,00 | — |

**Resultado esperado:** `realizado_nao_previsto`

**Observação:** Lançamento ocorreu no banco mas não havia previsão no fluxo. Motor deve listar como não conciliado no lado do extrato.

---

### CENARIO_05 — Valor diferente

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 12/06/2026 | 12/06/2026 |
| Razão Social / Categoria | CENARIO_05_VALOR_DIFERENTE RECURSOS LTDA | Folha de Pagamento (PJ) - CENARIO_05_VALOR_DIFERENTE |
| Valor | **-3.400,00** | **-3.600,00** |
| Diferença de valor | 200,00 (5,56%) | — |

**Resultado esperado:** `divergencia_valor`

**Tolerância de valor:** A diferença é de 5,56% — acima da tolerância de 1%. Deve gerar divergência de valor.

**Observação:** Folha paga com valor inferior ao previsto. O motor deve encontrar o par pela data e categoria, registrar a divergência de 200,00 e sinalizar para revisão manual.

---

### CENARIO_06 — Antecipação de recebimento

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data extrato | **12/06/2026** | — |
| Data fluxo | — | **15/06/2026** |
| Razão Social / Categoria | CENARIO_06_ANTECIPACAO_RECEBIMENTO CLIENTE SA | Receita Operacional - VENDAS (CLIENTE) - CENARIO_06 |
| Valor | +5.500,00 | +5.500,00 |
| Diferença de data | 3 dias (antecipado) | — |

**Resultado esperado:** `divergencia_data` (subtipo: antecipação)

**Observação:** Recebimento chegou 3 dias antes do previsto. Valor idêntico. Dentro da tolerância de data (3 dias). O motor deve conciliar com flag de divergência de data = antecipação.

---

### CENARIO_07 — Pagamento atrasado

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data extrato | **13/06/2026** | — |
| Data fluxo | — | **10/06/2026** |
| Razão Social / Categoria | CENARIO_07_PAGAMENTO_ATRASADO FORNECEDOR ME | Aluguel PE - CENARIO_07_PAGAMENTO_ATRASADO |
| Valor | -780,00 | -780,00 |
| Diferença de data | 3 dias (atrasado) | — |

**Resultado esperado:** `divergencia_data` (subtipo: atraso)

**Observação:** Pagamento ocorreu 3 dias após o previsto. Valor idêntico. Dentro da tolerância. O motor deve conciliar com flag de divergência de data = atraso.

---

### CENARIO_08 — Duplicidade no extrato

| Lançamento | Extrato | Fluxo de Caixa |
|---|---|---|
| 1ª ocorrência (CEN08a) | 13/06/2026, PIX ENVIADO, CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME, -200,00 | Limpeza / Diarista, 13/06/2026, -200,00 |
| 2ª ocorrência (CEN08b) | 13/06/2026, PIX ENVIADO, CENARIO_08_DUPLICIDADE_EXTRATO DIARISTA ME, -200,00 | — (sem correspondência) |

**Resultado esperado:**
- CEN08a → `conciliado`
- CEN08b → `duplicidade_extrato` (ou `realizado_nao_previsto`)

**Observação:** Mesmo beneficiário, mesma data, mesmo valor, dois lançamentos no extrato. O fluxo só tem uma previsão de -200,00. O motor deve conciliar um e sinalizar o outro como possível duplicidade. A decisão final (erro do banco vs. pagamento legítimo) é manual.

---

### CENARIO_09 — Duplicidade no fluxo de caixa

| Lançamento | Extrato | Fluxo de Caixa |
|---|---|---|
| Realizado (CEN09) | 14/06/2026, BOLETO PAGO, CENARIO_09_DUPLICIDADE_FLUXO SOFTWARE LTDA, -650,00 | — |
| Previsão A | — | Licenças de Software - CENARIO_09A_DUPLICIDADE_FLUXO, 14/06/2026, -650,00 |
| Previsão B | — | Licenças de Software - CENARIO_09B_DUPLICIDADE_FLUXO, 14/06/2026, -650,00 |

**Resultado esperado:**
- CEN09 (extrato) → `duplicidade_fluxo` — encontra dois candidatos no fluxo para o mesmo lançamento
- Previsão A → pode ser conciliada com CEN09, mas Previsão B fica `previsto_nao_realizado`

**Observação:** O fluxo previu o mesmo tipo de gasto duas vezes na mesma data. O extrato tem apenas um pagamento. O motor deve detectar a ambiguidade e sinalizar para revisão.

---

### CENARIO_10 — Pendente para análise manual

| Campo | Extrato | Fluxo de Caixa |
|---|---|---|
| Data | 14/06/2026 | — |
| Razão Social | CENARIO_10_PENDENTE_ANALISE TRANSF S/IDENTIFICACAO | — |
| Valor | +3.700,00 | — |

**Resultado esperado:** `pendente_analise_manual`

**Observação:** TED recebida sem identificação clara de origem. Não há correspondência direta no fluxo de caixa. O motor não deve tentar conciliar automaticamente — deve registrar como pendente para análise humana.

---

## Resumo dos resultados esperados

| Código | Cenário | Tipo esperado | E existe? | F existe? |
|---|---|---|---|---|
| CEN01a | Match perfeito (despesa) | conciliado | ✅ | ✅ |
| CEN01b | Match perfeito (receita) | conciliado | ✅ | ✅ |
| CEN01c | Match perfeito (adicional) | conciliado | ✅ | ✅ |
| CEN02  | Data diferente | divergencia_data | ✅ | ✅ |
| CEN03  | Previsto não realizado | previsto_nao_realizado | ❌ | ✅ |
| CEN04  | Realizado não previsto | realizado_nao_previsto | ✅ | ❌ |
| CEN05  | Valor diferente | divergencia_valor | ✅ | ✅ |
| CEN06  | Antecipação de recebimento | divergencia_data | ✅ | ✅ |
| CEN07  | Pagamento atrasado | divergencia_data | ✅ | ✅ |
| CEN08a | Duplicidade extrato (1ª) | conciliado | ✅ | ✅ |
| CEN08b | Duplicidade extrato (2ª) | duplicidade_extrato | ✅ | ❌ |
| CEN09  | Duplicidade fluxo | duplicidade_fluxo | ✅ | ✅×2 |
| CEN10  | Pendente análise manual | pendente_analise_manual | ✅ | ❌ |

---

## Tolerâncias iniciais recomendadas

| Tipo | Tolerância | Justificativa |
|---|---|---|
| Data | 3 dias corridos | Cobre fins de semana e feriados bancários |
| Valor | 1,00% do valor previsto | Cobre diferenças de centavos por arredondamento |
| Valor mínimo | R$ 0,01 | Diferenças menores são ignoradas |

---

## Contadores esperados após conciliação

| Status | Quantidade |
|---|---|
| Conciliados | 4 (CEN01a, CEN01b, CEN01c, CEN08a) |
| Divergência de data | 3 (CEN02, CEN06, CEN07) |
| Divergência de valor | 1 (CEN05) |
| Previsto não realizado | 2 (CEN03, CEN09B após resolução) |
| Realizado não previsto | 1 (CEN04) |
| Duplicidade no extrato | 1 (CEN08b) |
| Duplicidade no fluxo | 1 (CEN09) |
| Pendente análise manual | 1 (CEN10) |
| **Total de registros de extrato** | **13** |
| **Total de lançamentos de fluxo** | **11** |

---

## Como usar no smoke test do Storage

```bash
# 1. Subir os arquivos pela conciliação real no frontend:
#    /conciliacoes/48626b8e-cfb5-48d1-9774-a2c5607e9d55

# 2. Enviar na ordem:
#    TESTE_Extrato_Lancamentos_banco.xlsx  → tipo: extrato_bancario
#    TESTE_Fluxo_Caixa_Daxx.xlsx          → tipo: planilha_interna

# 3. Validar no banco:
SELECT nome_original, modo_retencao, arquivo_persistido, expira_em, hash_arquivo
FROM ia16_fechamento_dev.arquivos_enviados
ORDER BY criado_em DESC LIMIT 5;

# 4. Clicar em "Processar conciliação" e verificar status = processado
# 5. Confirmar logs em logs_processamento
```

---

## Próximo passo

Usar esta massa de teste como critério de aceite da Sprint 6.

O motor de conciliação deve gerar, para cada par de registros, um item em `itens_conciliacao` com:

```text
status: conciliado | divergente | pendente | duplicado | nao_encontrado | erro
```

E para cada divergência, um item em `divergencias_conciliacao` com:

```text
tipo: valor_diferente | data_diferente | registro_nao_encontrado | registro_duplicado | ...
```
"""

def criar_gabarito():
    caminho = SAIDA / "GABARITO_MASSA_TESTE_DAXX.md"
    caminho.write_text(GABARITO_MD, encoding="utf-8")
    print(f"  ✔ {caminho.name} ({caminho.stat().st_size // 1024} KB)")
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\nGerando massa de teste em: {SAIDA}\n")
    p1 = criar_extrato()
    p2 = criar_fluxo_caixa()
    p3 = criar_conciliacao_interna()
    p4 = criar_gabarito()

    # Verificar tamanhos
    print("\nVerificação de tamanhos:")
    for p in [p1, p2, p3]:
        kb = p.stat().st_size // 1024
        status = "✔ OK" if kb < 10240 else "✘ ACIMA DE 10MB"
        print(f"  {p.name}: {kb} KB — {status}")

    print(f"\nTotal de registros no extrato: {len(EXTRATO_LANCAMENTOS)}")
    print(f"Total de categorias no fluxo:  {len(FLUXO_CATEGORIAS)}")
    print(f"Cenários cobertos:             10 (CEN01–CEN10)\n")
