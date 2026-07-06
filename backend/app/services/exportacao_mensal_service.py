"""
Service de exportação da planilha mensal de conciliação.

Gera um arquivo .xlsx em memória com uma única aba no layout idêntico
ao export diário (exportacao_fechamento_service._gerar_excel_extrato_anotado):
mesmas larguras de coluna, alturas de linha, bordas, fontes, estilos,
saldo inicial/final, filtro automático e formatação de datas/números.

Para tipos bilaterais mantém o layout básico anterior.
Suporta qualquer tipo de conciliação (multicliente/multiconciliação).
"""
import io
import re
import calendar
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.arquivo_enviado import ArquivoEnviado
from app.models.empresa import Empresa
from app.models.fechamento_financeiro import FechamentoFinanceiro
from app.models.item_conciliacao import ItemConciliacao
from app.models.lancamento_extrato_anotado import LancamentoExtratoAnotado

# Status elegíveis para exportação
_STATUS_EXPORTAVEL = {"processado", "com_divergencias", "aprovado", "reaberto"}

_ABREV_MES = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
    5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

_COLUNAS_MENSAL = [
    "DATA",
    "DESCRIÇÃO LANÇAMENTO BANCO",
    "DESCRIÇÃO FORNECEDOR/CLIENTE",
    "NF / DOC",
    "VALOR NF/DOC",
    "ENTRADA EXTRATO",
    "SAIDA EXTRATO",
    "SALDO",
]

# Linha do cabeçalho da tabela (mesma do diário)
_LINHA_CABECALHO_TABELA = 8


# ── helpers ───────────────────────────────────────────────────────────────────

def _sanitizar_nome_arquivo(valor: str) -> str:
    return re.sub(r'[/\\:*?"<>|. ]', "_", valor.strip())


def _nome_aba(mes: int, ano: int) -> str:
    """Retorna nome da aba no formato 'Jun26'."""
    return f"{_ABREV_MES.get(mes, str(mes))}{str(ano)[-2:]}"


def _periodo_str_mensal(mes: int, ano: int) -> str:
    """Retorna 'Junho/2026'."""
    return f"{MESES_PT.get(mes, str(mes))}/{ano}"


def _saldo_inicial(lancamentos: list) -> float:
    """Calcula o saldo antes do primeiro lançamento (mesma lógica do serviço diário)."""
    if not lancamentos:
        return 0.0
    primeiro = lancamentos[0]
    if primeiro.saldo is None:
        return 0.0
    saldo_apos = float(primeiro.saldo)
    valor = float(primeiro.valor)
    return saldo_apos - valor if primeiro.tipo_movimento == "entrada" else saldo_apos + valor


# ── busca de metadados do arquivo de extrato ─────────────────────────────────

def _buscar_metadados_extrato(db: Session, ids_fechamentos: list) -> dict:
    """
    Busca os metadados mais completos dentre os arquivos de extrato do mês.

    Usa uma única query com .in_() (mesmo padrão do serviço diário).
    Percorre os arquivos em ordem cronológica e:
      1. Retorna imediatamente o primeiro que tiver agencia E conta.
      2. Guarda o primeiro que tiver pelo menos agencia OU conta como candidato.
      3. Se nenhum tiver metadados úteis, retorna {}.

    Critério: um arquivo é "útil" se tiver pelo menos agencia ou conta preenchidos.
    """
    arquivos = (
        db.query(ArquivoEnviado)
        .filter(
            ArquivoEnviado.fechamento_id.in_(ids_fechamentos),
            ArquivoEnviado.tipo_arquivo == "extrato_bancario",
        )
        .order_by(ArquivoEnviado.criado_em)
        .all()
    )
    melhor: dict = {}
    for arquivo in arquivos:
        meta = arquivo.metadados or {}
        agencia = meta.get("agencia", "")
        conta   = meta.get("conta", "")
        if agencia and conta:
            return meta          # Perfeito — tem ambos
        if (agencia or conta) and not melhor:
            melhor = meta        # Candidato parcial — continua procurando melhor
    return melhor


# ── busca de fechamentos ───────────────────────────────────────────────────────

def buscar_fechamentos_do_mes(
    db: Session,
    empresa_id,
    ano: int,
    mes: int,
    tipo_conciliacao: str,
    status_incluidos: set[str] | None = None,
) -> list[FechamentoFinanceiro]:
    """
    Retorna fechamentos da empresa, ano, mês e tipo informados,
    determinados pelo campo periodo_inicio, em ordem cronológica.
    """
    if status_incluidos is None:
        status_incluidos = _STATUS_EXPORTAVEL

    primeiro_dia = date(ano, mes, 1)
    ultimo_dia   = date(ano, mes, calendar.monthrange(ano, mes)[1])

    return (
        db.query(FechamentoFinanceiro)
        .filter(
            FechamentoFinanceiro.empresa_id     == empresa_id,
            FechamentoFinanceiro.tipo_conciliacao == tipo_conciliacao,
            FechamentoFinanceiro.status.in_(status_incluidos),
            FechamentoFinanceiro.periodo_inicio >= primeiro_dia,
            FechamentoFinanceiro.periodo_inicio <= ultimo_dia,
        )
        .order_by(FechamentoFinanceiro.periodo_inicio)
        .all()
    )


# ── layout idêntico ao diário (extrato_anotado) ───────────────────────────────

def _gerar_excel_extrato_anotado_mensal(
    empresa_nome: str,
    mes: int,
    ano: int,
    metadados: dict,
    lancamentos: list,
) -> tuple[bytes, str]:
    """
    Gera o Excel mensal com exatamente o mesmo layout visual do export diário:
    mesmas larguras, alturas, bordas, fontes, estilos, saldo inicial/final,
    filtro automático, formatação de datas e números.
    """
    # ── Estilos (idênticos ao serviço diário) ────────────────────────────────
    fill_azul    = PatternFill("solid", fgColor="1F4E79")
    fill_periodo = PatternFill("solid", fgColor="BDD7EE")
    fill_saldo   = PatternFill("solid", fgColor="E2EFDA")
    font_h       = Font(bold=True, color="FFFFFF", size=14)
    font_b12     = Font(bold=True, size=12)
    font_b11     = Font(bold=True, size=11)
    font_n11     = Font(size=11)
    font_n12     = Font(size=12)
    font_verde   = Font(size=11, color="00B050")
    font_vermelho= Font(size=11, color="FF0000")
    aln_c        = Alignment(horizontal="center", vertical="center")
    aln_r        = Alignment(horizontal="right",  vertical="center")
    aln_l        = Alignment(horizontal="left",   vertical="center")
    lado_fino    = Side(style="thin", color="BFBFBF")
    borda        = Border(left=lado_fino, right=lado_fino, top=lado_fino, bottom=lado_fino)
    FMT_NUM      = "#,##0.00"
    FMT_DATA     = "DD/MM/YYYY"

    # ── Metadados do extrato ──────────────────────────────────────────────────
    atualizacao = metadados.get("atualizacao", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    nome_banco  = metadados.get("nome", empresa_nome)
    agencia     = metadados.get("agencia", "")
    conta       = metadados.get("conta", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _nome_aba(mes, ano)

    # ── Larguras das colunas (idênticas ao diário) ────────────────────────────
    ws.column_dimensions["A"].width = 21.1
    ws.column_dimensions["B"].width = 36.9
    ws.column_dimensions["C"].width = 79.6
    ws.column_dimensions["D"].width = 20.0
    ws.column_dimensions["E"].width = 15.6
    ws.column_dimensions["F"].width = 16.4
    ws.column_dimensions["G"].width = 15.6
    ws.column_dimensions["H"].width = 18.0

    # ── L1-4: metadados do banco ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 15.6
    ws.row_dimensions[2].height = 15.6
    ws.row_dimensions[3].height = 15.6
    ws.row_dimensions[4].height = 15.6

    ws.cell(1, 1, "Atualização:").font = font_n12
    c = ws.cell(1, 2, atualizacao); c.font = font_b12
    ws.merge_cells("B1:D1")

    ws.cell(2, 1, "Nome:").font = font_n12
    ws.cell(2, 2, nome_banco).font = font_b12

    ws.cell(3, 1, "Agência:").font = font_n12
    ws.cell(3, 2, agencia).font = font_b12

    ws.cell(4, 1, "Conta:").font = font_n12
    c = ws.cell(4, 2, conta); c.font = font_b12
    ws.merge_cells("B4:D4")

    # ── L5: espaçador fino ───────────────────────────────────────────────────
    ws.row_dimensions[5].height = 9.6

    # ── L6: Período ──────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 15.6
    c = ws.cell(6, 1, f"Periodo:  {_periodo_str_mensal(mes, ano)}")
    c.font = font_n12
    c.fill = fill_periodo

    # ── L7: espaçador muito fino ─────────────────────────────────────────────
    ws.row_dimensions[7].height = 5.4

    # ── L8: cabeçalho da tabela ───────────────────────────────────────────────
    ROW_HEADER = 8
    ws.row_dimensions[ROW_HEADER].height = 37.2
    for col, label in enumerate(_COLUNAS_MENSAL, start=1):
        c = ws.cell(ROW_HEADER, col, label)
        c.fill      = fill_azul
        c.font      = font_h
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
    ws.auto_filter.ref = f"A{ROW_HEADER}:H{ROW_HEADER}"

    # ── L9: saldo inicial ─────────────────────────────────────────────────────
    ROW_SALDO = ROW_HEADER + 1
    saldo_ini = _saldo_inicial(lancamentos)
    for col in range(1, 9):
        c = ws.cell(ROW_SALDO, col)
        c.border = borda
        c.font   = font_b11
        c.fill   = fill_saldo
    ws.cell(ROW_SALDO, 1, lancamentos[0].data_lancamento if lancamentos else None)
    ws.cell(ROW_SALDO, 1).number_format = FMT_DATA
    ws.cell(ROW_SALDO, 2, "SALDO TOTAL DISPONÍVEL DIA")
    ws.merge_cells(f"B{ROW_SALDO}:G{ROW_SALDO}")
    c = ws.cell(ROW_SALDO, 8, saldo_ini)
    c.number_format = FMT_NUM
    c.alignment     = aln_r
    c.font          = font_b11
    c.fill          = fill_saldo

    # ── L10+: linhas de dados ─────────────────────────────────────────────────
    ROW_INICIO = ROW_SALDO + 1
    for i, l in enumerate(lancamentos):
        row  = ROW_INICIO + i
        prev = row - 1

        entrada = float(l.valor) if l.tipo_movimento == "entrada" else None
        saida   = float(l.valor) if l.tipo_movimento == "saida"   else None
        nf_val  = float(l.valor_nf_doc) if l.valor_nf_doc is not None else None

        # DATA
        c = ws.cell(row, 1, l.data_lancamento)
        c.number_format = FMT_DATA; c.border = borda; c.font = font_n11; c.alignment = aln_c

        # DESCRIÇÃO BANCO
        c = ws.cell(row, 2, l.descricao_banco or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_l

        # DESCRIÇÃO FORNECEDOR/CLIENTE
        c = ws.cell(row, 3, l.descricao_negocio or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_l

        # NF/DOC
        c = ws.cell(row, 4, l.nf_doc or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_c

        # VALOR NF/DOC
        c = ws.cell(row, 5, nf_val)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_n11; c.alignment = aln_r

        # ENTRADA (verde)
        c = ws.cell(row, 6, entrada)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_verde; c.alignment = aln_r

        # SAÍDA (vermelho)
        c = ws.cell(row, 7, saida)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_vermelho; c.alignment = aln_r

        # SALDO (fórmula acumulada)
        c = ws.cell(row, 8, f"=H{prev}+F{row}-G{row}")
        c.number_format = FMT_NUM; c.border = borda; c.font = font_n11; c.alignment = aln_r

    # ── Linha final: saldo total disponível ───────────────────────────────────
    last_row = ROW_INICIO + len(lancamentos) - 1
    row_fim  = last_row + 1
    for col in range(1, 9):
        c = ws.cell(row_fim, col)
        c.border = borda; c.font = font_b11; c.fill = fill_saldo
    ws.cell(row_fim, 1, lancamentos[-1].data_lancamento if lancamentos else None)
    ws.cell(row_fim, 1).number_format = FMT_DATA
    ws.cell(row_fim, 2, "SALDO TOTAL DISPONÍVEL DIA")
    ws.merge_cells(f"B{row_fim}:G{row_fim}")
    c = ws.cell(row_fim, 8, f"=H{last_row}")
    c.number_format = FMT_NUM; c.alignment = aln_r; c.font = font_b11; c.fill = fill_saldo

    # ── Salvar ────────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_mes    = MESES_PT.get(mes, str(mes))
    empresa_slug = _sanitizar_nome_arquivo(empresa_nome)
    return buffer.read(), f"Conciliacao_{empresa_slug}_{nome_mes}_{ano}.xlsx"


# ── layout básico para tipos bilaterais ───────────────────────────────────────

def _gerar_excel_bilateral_mensal(
    empresa_nome: str,
    mes: int,
    ano: int,
    metadados: dict,
    itens: list,
) -> tuple[bytes, str]:
    """
    Layout para tipos bilaterais (sem semântica de saldo/entrada/saída).
    Mantém o estilo básico anterior com cabeçalho azul, período e filtro.
    """
    fill_azul    = PatternFill("solid", fgColor="1F4E79")
    fill_periodo = PatternFill("solid", fgColor="BDD7EE")
    font_h       = Font(bold=True, color="FFFFFF")
    font_label   = Font(bold=True)
    aln_c        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    FMT_NUM      = "#,##0.00"
    FMT_DATA     = "DD/MM/YYYY"

    atualizacao = metadados.get("atualizacao", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    nome_banco  = metadados.get("nome", empresa_nome)
    agencia     = metadados.get("agencia", "")
    conta       = metadados.get("conta", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _nome_aba(mes, ano)

    ws.cell(1, 1, "Atualização:").font = font_label
    ws.cell(1, 2, atualizacao)
    ws.cell(2, 1, "Nome:").font = font_label
    ws.cell(2, 2, nome_banco)
    ws.cell(3, 1, "Agência:").font = font_label
    ws.cell(3, 2, agencia)
    ws.cell(4, 1, "Conta:").font = font_label
    ws.cell(4, 2, conta)
    c = ws.cell(6, 1, f"Periodo:  {_periodo_str_mensal(mes, ano)}")
    c.font = font_label
    c.fill = fill_periodo

    for col, label in enumerate(_COLUNAS_MENSAL, start=1):
        c = ws.cell(8, col, label)
        c.fill = fill_azul
        c.font = font_h
        c.alignment = aln_c
    ws.auto_filter.ref = "A8:H8"

    inicio = 9
    for i, item in enumerate(itens):
        row    = inicio + i
        data   = item.data_realizada or item.data_prevista
        entrada = float(item.valor_realizado) if getattr(item, "tipo_movimento", None) == "entrada" else None
        saida   = float(item.valor_realizado) if getattr(item, "tipo_movimento", None) == "saida"   else None
        c = ws.cell(row, 1, data); c.number_format = FMT_DATA
        ws.cell(row, 2, item.descricao_realizada or item.descricao_prevista or "")
        ws.cell(row, 3, item.descricao_prevista or "")
        c = ws.cell(row, 6, entrada); c.number_format = FMT_NUM
        c = ws.cell(row, 7, saida);   c.number_format = FMT_NUM

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_mes     = MESES_PT.get(mes, str(mes))
    empresa_slug = _sanitizar_nome_arquivo(empresa_nome)
    return buffer.read(), f"Conciliacao_{empresa_slug}_{nome_mes}_{ano}.xlsx"


# ── função principal ───────────────────────────────────────────────────────────

def gerar_excel_mensal(
    db: Session,
    empresa_id,
    ano: int,
    mes: int,
    tipo_conciliacao: str,
    status_incluidos: set[str] | None = None,
) -> tuple[bytes, str]:
    """
    Gera a planilha mensal de conciliação para uma empresa, tipo e mês.

    Para extrato_anotado: layout idêntico ao export diário.
    Para outros tipos: layout básico bilateral.

    Retorna (bytes_do_arquivo, nome_do_arquivo).
    Levanta ValueError se não houver fechamentos no período.
    """
    fechamentos = buscar_fechamentos_do_mes(
        db, empresa_id, ano, mes, tipo_conciliacao, status_incluidos
    )

    if not fechamentos:
        raise ValueError(
            f"Nenhuma conciliação encontrada para {MESES_PT.get(mes, str(mes))}/{ano} "
            f"com tipo '{tipo_conciliacao}'."
        )

    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    empresa_nome = empresa.nome if empresa else str(empresa_id)

    ids_fechamentos = [f.id for f in fechamentos]
    metadados       = _buscar_metadados_extrato(db, ids_fechamentos)

    if tipo_conciliacao == "extrato_anotado":
        lancamentos = (
            db.query(LancamentoExtratoAnotado)
            .filter(LancamentoExtratoAnotado.fechamento_id.in_(ids_fechamentos))
            .order_by(
                LancamentoExtratoAnotado.data_lancamento,
                LancamentoExtratoAnotado.fechamento_id,
                LancamentoExtratoAnotado.linha_origem,
            )
            .all()
        )
        return _gerar_excel_extrato_anotado_mensal(empresa_nome, mes, ano, metadados, lancamentos)

    itens = (
        db.query(ItemConciliacao)
        .filter(ItemConciliacao.fechamento_id.in_(ids_fechamentos))
        .order_by(ItemConciliacao.data_prevista, ItemConciliacao.fechamento_id)
        .all()
    )
    return _gerar_excel_bilateral_mensal(empresa_nome, mes, ano, metadados, itens)
