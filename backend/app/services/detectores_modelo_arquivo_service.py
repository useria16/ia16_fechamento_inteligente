"""
Serviço de detecção de modelo de arquivo.

Dado um arquivo .xlsx e um tipo de arquivo (extrato_bancario, planilha_interna, etc.),
identifica qual modelo cadastrado melhor descreve a estrutura do arquivo.

Estratégia:
  - Para cada modelo candidato (mesmo tipo_arquivo, ativo=True), calcula confiança
  - Detecção por estrutura, não por nome de cliente
  - Retorna o modelo com maior confiança acima do limiar (LIMIAR_DETECCAO)
"""
import io
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string
from sqlalchemy.orm import Session

from app.models.modelo_arquivo import ModeloArquivo
from app.schemas.normalizacao import ResultadoDeteccao

LIMIAR_DETECCAO = 0.60  # confiança mínima para aceitar detecção


def detectar_modelo(
    conteudo: bytes,
    tipo_arquivo: str,
    db: Session,
) -> ResultadoDeteccao:
    """
    Identifica o modelo de arquivo mais compatível com o conteúdo fornecido.

    Busca modelos globais (empresa_id IS NULL) e ativos com o tipo_arquivo informado.
    """
    modelos = (
        db.query(ModeloArquivo)
        .filter(
            ModeloArquivo.tipo_arquivo == tipo_arquivo,
            ModeloArquivo.ativo == True,  # noqa: E712
        )
        .all()
    )

    melhor_modelo = None
    melhor_confianca = 0.0
    melhor_motivos: list[str] = []

    for modelo in modelos:
        if not modelo.tipo_estrutura:
            continue
        confianca, motivos = _calcular_confianca(conteudo, modelo)
        if confianca > melhor_confianca:
            melhor_confianca = confianca
            melhor_modelo = modelo
            melhor_motivos = motivos

    if melhor_modelo and melhor_confianca >= LIMIAR_DETECCAO:
        return ResultadoDeteccao(
            detectado=True,
            modelo_arquivo_id=str(melhor_modelo.id),
            codigo_modelo=melhor_modelo.codigo,
            confianca=melhor_confianca,
            motivos=melhor_motivos,
        )

    return ResultadoDeteccao(
        detectado=False,
        modelo_arquivo_id=None,
        codigo_modelo=None,
        confianca=melhor_confianca,
        motivos=melhor_motivos or ["Nenhum modelo compatível encontrado"],
    )


def _calcular_confianca(
    conteudo: bytes,
    modelo: ModeloArquivo,
) -> tuple[float, list[str]]:
    config = modelo.mapeamento_colunas or {}
    try:
        if modelo.tipo_estrutura == "tabular":
            return _detectar_tabular(conteudo, config)
        if modelo.tipo_estrutura == "transposto":
            return _detectar_transposto(conteudo, config)
    except Exception as exc:
        return 0.0, [f"Erro durante detecção: {exc}"]
    return 0.0, ["Tipo de estrutura não suportado para detecção"]


def _detectar_tabular(conteudo: bytes, config: dict) -> tuple[float, list[str]]:
    """
    Detecta extrato tabular verificando:
    1. Aba correta existe no arquivo
    2. Cabeçalho encontrado na linha esperada
    3. Maioria das colunas esperadas presente
    4. Dados existem na linha de início
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    motivos: list[str] = []
    pontos = 0.0
    total = 4.0

    aba_nome = config.get("aba")
    linha_cab = config.get("linha_cabecalho", 10)
    colunas_esperadas = config.get("colunas_esperadas", [
        "Data", "Lançamento", "Razão Social", "CPF/CNPJ", "Valor (R$)", "Saldo (R$)",
    ])
    linha_inicio = config.get("linha_inicio_dados", linha_cab + 1)

    # Check 1: aba existe (obrigatório — sem aba, sem detecção)
    if aba_nome and aba_nome not in wb.sheetnames:
        wb.close()
        motivos.append(f"Aba '{aba_nome}' não encontrada (abas: {wb.sheetnames})")
        return 0.0, motivos

    ws = wb[aba_nome] if aba_nome else wb.active
    pontos += 1.0
    motivos.append(f"Aba '{aba_nome}' encontrada")

    # Check 2: cabeçalho na linha esperada
    cabecalho: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 == linha_cab:
            cabecalho = [str(c).strip() if c else "" for c in row]
            break

    if cabecalho:
        pontos += 0.5
        motivos.append(f"Conteúdo encontrado na linha {linha_cab}")
    else:
        motivos.append(f"Linha {linha_cab} vazia ou inexistente")
        wb.close()
        return round(pontos / total, 2), motivos

    # Check 3: colunas esperadas
    colunas_ok = sum(1 for ce in colunas_esperadas if ce in cabecalho)
    taxa = colunas_ok / max(len(colunas_esperadas), 1)
    pontos += taxa * 1.5  # peso maior para colunas corretas
    motivos.append(f"{colunas_ok}/{len(colunas_esperadas)} colunas esperadas encontradas")

    # Check 4: dados na linha de início
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 == linha_inicio:
            if any(c for c in row if c is not None):
                pontos += 1.0
                motivos.append(f"Dados encontrados na linha {linha_inicio}")
            else:
                motivos.append(f"Linha {linha_inicio} vazia")
            break

    wb.close()
    return round(min(pontos / total, 1.0), 2), motivos


def _detectar_transposto(conteudo: bytes, config: dict) -> tuple[float, list[str]]:
    """
    Detecta fluxo transposto verificando:
    1. Aba correta existe (crédito parcial se não encontrar)
    2. Linha de datas tem datas reais nas colunas a partir de D
    3. Coluna B tem categorias de texto
    4. Dados (valores não-zero) existem no corpo da planilha
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    motivos: list[str] = []
    pontos = 0.0
    total = 4.0

    aba_nome = config.get("aba")
    linha_datas = config.get("linha_datas", 1)
    col_inicio_idx = column_index_from_string(config.get("coluna_inicio_valores", "D"))
    col_cat_idx = column_index_from_string(config.get("coluna_categoria", "B"))

    # Check 1: aba (crédito parcial se não encontrar)
    if aba_nome and aba_nome in wb.sheetnames:
        ws = wb[aba_nome]
        pontos += 1.0
        motivos.append(f"Aba '{aba_nome}' encontrada")
    else:
        ws = wb.active
        if aba_nome:
            pontos += 0.3
            motivos.append(f"Aba '{aba_nome}' não encontrada, usando aba ativa '{ws.title}'")
        else:
            pontos += 0.5
            motivos.append(f"Usando aba ativa '{ws.title}'")

    # Check 2: datas na linha 1 a partir de col D
    date_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 != linha_datas:
            continue
        for col_idx, val in enumerate(row, 1):
            if col_idx >= col_inicio_idx and isinstance(val, (date, datetime)):
                date_count += 1
        break

    if date_count >= 5:
        pontos += 1.0
        motivos.append(f"{date_count} datas encontradas na linha {linha_datas} a partir da coluna D")
    elif date_count >= 2:
        pontos += 0.5
        motivos.append(f"Apenas {date_count} datas encontradas (esperado >= 5)")
    else:
        motivos.append(f"Datas insuficientes na linha {linha_datas} ({date_count} encontradas)")

    # Check 3: categorias na coluna B
    cat_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 <= linha_datas:
            continue
        if i > linha_datas + 30:
            break
        if len(row) >= col_cat_idx:
            val = row[col_cat_idx - 1]
            if isinstance(val, str) and val.strip():
                cat_count += 1

    if cat_count >= 5:
        pontos += 1.0
        motivos.append(f"{cat_count} categorias de texto encontradas na coluna B")
    elif cat_count >= 2:
        pontos += 0.5
        motivos.append(f"Poucas categorias na coluna B ({cat_count})")
    else:
        motivos.append(f"Categorias insuficientes na coluna B ({cat_count})")

    # Check 4: valores no corpo da planilha
    valor_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 <= linha_datas:
            continue
        if i > linha_datas + 30:
            break
        for col_idx, val in enumerate(row, 1):
            if col_idx >= col_inicio_idx and isinstance(val, (int, float)) and val != 0:
                valor_count += 1
        if valor_count >= 3:
            break

    if valor_count >= 3:
        pontos += 1.0
        motivos.append(f"Valores numéricos encontrados na área de dados")
    else:
        motivos.append("Poucos valores na área de dados")

    wb.close()
    return round(min(pontos / total, 1.0), 2), motivos
