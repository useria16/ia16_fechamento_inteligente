"""
Service de exportação do pacote final do fechamento em Excel.

Gera um arquivo .xlsx em memória com 5 abas:
  1. Resumo
  2. Itens Conciliados
  3. Divergências
  4. Pendências
  5. Observações da Revisão
"""
import io
import re
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.divergencia_conciliacao import DivergenciaConciliacao
from app.models.empresa import Empresa
from app.models.fechamento_financeiro import FechamentoFinanceiro
from app.models.item_conciliacao import ItemConciliacao
from app.models.arquivo_enviado import ArquivoEnviado
from app.models.lancamento_extrato_anotado import LancamentoExtratoAnotado
from app.models.usuario import Usuario


# ── helpers ────────────────────────────────────────────────────────────────────

def _sanitizar_nome(valor: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_-]", "_", valor.strip().lower())


def _estilizar_cabecalho(ws, linha: int, num_colunas: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=linha, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _ajustar_largura(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _fmt(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d %H:%M")
    return str(valor)


# ── aba Resumo ─────────────────────────────────────────────────────────────────

def _aba_resumo(ws, fechamento: FechamentoFinanceiro, empresa_nome: str, aprovado_por: str | None, reaberto_por: str | None) -> None:
    ws.title = "Resumo"
    campos = [
        ("Empresa", empresa_nome),
        ("Título da Conciliação", fechamento.titulo),
        ("Tipo de Conciliação", fechamento.tipo_conciliacao),
        ("Status", fechamento.status),
        ("Período Início", _fmt(fechamento.periodo_inicio)),
        ("Período Fim", _fmt(fechamento.periodo_fim)),
        ("Data de Criação", _fmt(fechamento.criado_em)),
        ("Data de Processamento", _fmt(fechamento.atualizado_em)),
        ("Data de Aprovação", _fmt(fechamento.aprovado_em)),
        ("Aprovado por", aprovado_por or ""),
        ("Observação de Aprovação", fechamento.observacao_aprovacao or ""),
        ("Motivo da Reabertura", fechamento.motivo_reabertura or ""),
        ("Reaberto em", _fmt(fechamento.reaberto_em)),
        ("Reaberto por", reaberto_por or ""),
        ("Quantidade de Registros", fechamento.quantidade_registros),
        ("Quantidade de Conciliados", fechamento.quantidade_conciliados),
        ("Quantidade de Divergentes", fechamento.quantidade_divergentes),
        ("Quantidade de Pendentes", fechamento.quantidade_pendentes),
        ("Valor Total Processado", float(fechamento.valor_total_processado or 0)),
        ("Valor Total Conciliado", float(fechamento.valor_total_conciliado or 0)),
        ("Valor Total Divergente", float(fechamento.valor_total_divergente or 0)),
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    for i, (campo, valor) in enumerate(campos, start=1):
        cell_a = ws.cell(row=i, column=1, value=campo)
        cell_a.fill = header_fill
        cell_a.font = header_font
        ws.cell(row=i, column=2, value=valor)


# ── aba Itens Conciliados ──────────────────────────────────────────────────────

def _aba_itens_conciliados(ws, itens: list[ItemConciliacao]) -> None:
    ws.title = "Itens Conciliados"
    cabecalho = [
        "Data Prevista", "Data Realizada", "Descrição Prevista", "Descrição Realizada",
        "Tipo de Movimento", "Valor Previsto", "Valor Realizado",
        "Diferença de Valor", "Diferença em Dias", "Confiança",
    ]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for item in itens:
        ws.append([
            _fmt(item.data_prevista),
            _fmt(item.data_realizada),
            item.descricao_prevista or "",
            item.descricao_realizada or "",
            item.tipo_movimento,
            float(item.valor_previsto) if item.valor_previsto is not None else "",
            float(item.valor_realizado) if item.valor_realizado is not None else "",
            float(item.diferenca_valor) if item.diferenca_valor is not None else "",
            item.diferenca_dias if item.diferenca_dias is not None else "",
            float(item.confianca) if item.confianca is not None else "",
        ])
    _ajustar_largura(ws)


# ── aba Divergências ───────────────────────────────────────────────────────────

def _aba_divergencias(ws, divergencias: list[DivergenciaConciliacao]) -> None:
    ws.title = "Divergências"
    cabecalho = [
        "Tipo de Divergência", "Severidade", "Status da Revisão", "Descrição",
        "Valor Previsto", "Valor Realizado", "Diferença de Valor",
        "Data Prevista", "Data Realizada", "Diferença em Dias",
        "Observação", "Resolvido em", "Atualizado em",
    ]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for d in divergencias:
        ws.append([
            d.tipo_divergencia,
            d.severidade,
            d.status,
            d.descricao,
            float(d.valor_previsto) if d.valor_previsto is not None else "",
            float(d.valor_realizado) if d.valor_realizado is not None else "",
            float(d.diferenca_valor) if d.diferenca_valor is not None else "",
            _fmt(d.data_prevista),
            _fmt(d.data_realizada),
            d.diferenca_dias if d.diferenca_dias is not None else "",
            d.observacao or "",
            _fmt(d.resolvido_em),
            _fmt(d.atualizado_em),
        ])
    _ajustar_largura(ws)


# ── aba Pendências ─────────────────────────────────────────────────────────────

def _aba_pendencias(ws, pendencias: list[DivergenciaConciliacao]) -> None:
    ws.title = "Pendências"
    cabecalho = [
        "Tipo", "Severidade", "Status da Revisão", "Descrição",
        "Valor Previsto", "Valor Realizado",
        "Data Prevista", "Data Realizada",
        "Observação", "Resolvido em", "Atualizado em",
    ]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for p in pendencias:
        ws.append([
            p.tipo_divergencia,
            p.severidade,
            p.status,
            p.descricao,
            float(p.valor_previsto) if p.valor_previsto is not None else "",
            float(p.valor_realizado) if p.valor_realizado is not None else "",
            _fmt(p.data_prevista),
            _fmt(p.data_realizada),
            p.observacao or "",
            _fmt(p.resolvido_em),
            _fmt(p.atualizado_em),
        ])
    _ajustar_largura(ws)


# ── aba Observações da Revisão ─────────────────────────────────────────────────

def _aba_observacoes(ws, divergencias_com_obs: list[DivergenciaConciliacao]) -> None:
    ws.title = "Observações da Revisão"
    cabecalho = [
        "Tipo de Divergência", "Status da Revisão", "Observação",
        "Atualizado por (ID)", "Atualizado em", "Resolvido em",
    ]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for d in divergencias_com_obs:
        ws.append([
            d.tipo_divergencia,
            d.status,
            d.observacao or "",
            str(d.atualizado_por_usuario_id) if d.atualizado_por_usuario_id else "",
            _fmt(d.atualizado_em),
            _fmt(d.resolvido_em),
        ])
    _ajustar_largura(ws)


# ── aba Extrato Anotado ────────────────────────────────────────────────────────

_LABELS_CONFERENCIA = {
    "encontrado":             "Previsto no fluxo",
    "data_diferente":         "Data diferente",
    "nao_encontrado":         "Não encontrado",
    "valor_diferente":        "Valor diferente",
    "possivel_correspondencia": "Revisar",
    "pendente_analise":       "Pendente",
}


def _aba_extrato_anotado(ws, lancamentos: list[LancamentoExtratoAnotado]) -> None:
    ws.title = "Extrato Anotado"
    tem_conferencia = any(l.tipo_conferencia_fluxo is not None for l in lancamentos)
    cabecalho = [
        "DATA", "DESCRIÇÃO LANÇAMENTO BANCO", "DESCRIÇÃO FORNECEDOR/CLIENTE",
        "NF / DOC", "VALOR NF/DOC",
        "ENTRADA EXTRATO", "SAÍDA EXTRATO", "SALDO",
        "OBSERVAÇÃO", "CATEGORIA", "STATUS REVISÃO",
    ]
    if tem_conferencia:
        cabecalho.append("STATUS NO FLUXO")

    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for l in lancamentos:
        entrada = float(l.valor) if l.tipo_movimento == "entrada" else ""
        saida   = float(l.valor) if l.tipo_movimento == "saida"   else ""
        saldo   = float(l.saldo) if l.saldo is not None else ""
        linha = [
            _fmt(l.data_lancamento),
            l.descricao_banco or "",
            l.descricao_negocio or "",
            l.nf_doc or "",
            float(l.valor_nf_doc) if l.valor_nf_doc is not None else "",
            entrada,
            saida,
            saldo,
            l.observacao or "",
            l.categoria or l.categoria_sugerida or "",
            l.status_revisao,
        ]
        if tem_conferencia:
            linha.append(_LABELS_CONFERENCIA.get(l.tipo_conferencia_fluxo or "", ""))
        ws.append(linha)
    _ajustar_largura(ws)


def _aba_conferencia_fluxo(ws, lancamentos: list[LancamentoExtratoAnotado]) -> None:
    ws.title = "Conferência com Fluxo"
    cabecalho = [
        "DATA EXTRATO", "DESCRIÇÃO BANCO", "TIPO MOVIMENTO",
        "VALOR EXTRATO", "STATUS NO FLUXO",
        "DATA PREVISTA", "VALOR PREVISTO", "DESCRIÇÃO PREVISTA",
        "OBSERVAÇÃO SISTEMA", "CONFIANÇA",
    ]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))

    for l in lancamentos:
        ws.append([
            _fmt(l.data_lancamento),
            l.descricao_banco or "",
            l.tipo_movimento,
            float(l.valor),
            _LABELS_CONFERENCIA.get(l.tipo_conferencia_fluxo or "", l.tipo_conferencia_fluxo or ""),
            _fmt(l.data_prevista) if l.data_prevista else "",
            float(l.valor_previsto) if l.valor_previsto is not None else "",
            l.descricao_prevista or "",
            l.observacao_sistema or "",
            float(l.confianca_conferencia) if l.confianca_conferencia is not None else "",
        ])
    _ajustar_largura(ws)


def _aba_pendencias_extrato(ws, pendentes: list[LancamentoExtratoAnotado]) -> None:
    ws.title = "Pendências de Revisão"
    cabecalho = ["DATA", "DESCRIÇÃO BANCO", "RAZÃO SOCIAL", "VALOR", "TIPO", "CATEGORIA SUGERIDA", "STATUS"]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))
    for l in pendentes:
        ws.append([
            _fmt(l.data_lancamento), l.descricao_banco or "", l.razao_social or "",
            float(l.valor), l.tipo_movimento,
            l.categoria_sugerida or "", l.status_revisao,
        ])
    _ajustar_largura(ws)


def _aba_observacoes_extrato(ws, com_obs: list[LancamentoExtratoAnotado]) -> None:
    ws.title = "Observações"
    cabecalho = ["DATA", "DESCRIÇÃO BANCO", "CATEGORIA", "OBSERVAÇÃO", "ATUALIZADO EM", "STATUS"]
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, 1, len(cabecalho))
    for l in com_obs:
        ws.append([
            _fmt(l.data_lancamento), l.descricao_banco or "",
            l.categoria or l.categoria_sugerida or "",
            l.observacao or "", _fmt(l.atualizado_em), l.status_revisao,
        ])
    _ajustar_largura(ws)


# ── função principal ───────────────────────────────────────────────────────────

def gerar_excel_fechamento(db: Session, fechamento: FechamentoFinanceiro) -> tuple[bytes, str]:
    if fechamento.tipo_conciliacao == "extrato_anotado":
        return _gerar_excel_extrato_anotado(db, fechamento)
    return _gerar_excel_bilateral(db, fechamento)


def _periodo_label(fechamento: FechamentoFinanceiro) -> str:
    """Retorna label curto do período para nome da aba, ex: 'Jun26'."""
    meses = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
             7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
    d = fechamento.periodo_inicio
    if d and hasattr(d, "month"):
        return f"{meses.get(d.month, str(d.month))}{str(d.year)[2:]}"
    return "Extrato"


def _periodo_str(fechamento: FechamentoFinanceiro) -> str:
    """Retorna string legível do período, ex: 'Junho/2026'."""
    meses = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
             7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
    d = fechamento.periodo_inicio
    if d and hasattr(d, "month"):
        return f"{meses.get(d.month, str(d.month))}/{d.year}"
    return ""


def _saldo_inicial(lancamentos: list[LancamentoExtratoAnotado]) -> float:
    """Calcula o saldo antes do primeiro lançamento."""
    if not lancamentos:
        return 0.0
    primeiro = lancamentos[0]
    if primeiro.saldo is None:
        return 0.0
    saldo_apos = float(primeiro.saldo)
    valor = float(primeiro.valor)
    return saldo_apos - valor if primeiro.tipo_movimento == "entrada" else saldo_apos + valor


def _borda_fina() -> Border:
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _gerar_excel_extrato_anotado(db: Session, fechamento: FechamentoFinanceiro) -> tuple[bytes, str]:
    empresa = db.query(Empresa).filter(Empresa.id == fechamento.empresa_id).first()
    empresa_nome = empresa.nome if empresa else str(fechamento.empresa_id)

    lancamentos = (
        db.query(LancamentoExtratoAnotado)
        .filter(LancamentoExtratoAnotado.fechamento_id == fechamento.id)
        .order_by(LancamentoExtratoAnotado.data_lancamento, LancamentoExtratoAnotado.linha_origem)
        .all()
    )

    # Metadados do banco (Agência, Conta, Nome) gravados no upload
    arquivo_extrato = (
        db.query(ArquivoEnviado)
        .filter(
            ArquivoEnviado.fechamento_id == fechamento.id,
            ArquivoEnviado.tipo_arquivo == "extrato_bancario",
        )
        .first()
    )
    meta_banco = (arquivo_extrato.metadados or {}) if arquivo_extrato else {}
    nome_banco  = meta_banco.get("nome", empresa_nome)
    agencia     = meta_banco.get("agencia", "")
    conta       = meta_banco.get("conta", "")
    atualizacao = meta_banco.get("atualizacao", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

    wb = openpyxl.Workbook()
    ws = wb.active  # type: ignore[assignment]
    ws.title = _periodo_label(fechamento)

    # ── Estilos ───────────────────────────────────────────────────────────────
    fill_azul    = PatternFill("solid", fgColor="1F4E79")
    fill_periodo = PatternFill("solid", fgColor="BDD7EE")
    fill_saldo   = PatternFill("solid", fgColor="E2EFDA")
    fill_total   = PatternFill("solid", fgColor="D9E1F2")
    font_h       = Font(bold=True, color="FFFFFF", size=14)
    font_b12     = Font(bold=True, size=12)
    font_b11     = Font(bold=True, size=11)
    font_n11     = Font(size=11)
    font_n12     = Font(size=12)
    font_verde   = Font(size=11, color="00B050")
    font_vermelho= Font(size=11, color="FF0000")
    aln_c        = Alignment(horizontal="center", vertical="center")
    aln_r        = Alignment(horizontal="right",  vertical="center")
    aln_l        = Alignment(horizontal="left",   vertical="center")
    lado_fino    = Side(style="thin", color="BFBFBF")
    borda        = Border(left=lado_fino, right=lado_fino, top=lado_fino, bottom=lado_fino)
    FMT_NUM      = "#,##0.00"
    FMT_DATA     = "DD/MM/YYYY"

    # ── Larguras das colunas ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 21.1
    ws.column_dimensions["B"].width = 36.9
    ws.column_dimensions["C"].width = 79.6
    ws.column_dimensions["D"].width = 20.0
    ws.column_dimensions["E"].width = 15.6
    ws.column_dimensions["F"].width = 16.4
    ws.column_dimensions["G"].width = 15.6
    ws.column_dimensions["H"].width = 18.0

    # ── L1-4: metadados do banco (sem linhas vazias antes) ───────────────────
    ws.row_dimensions[1].height = 15.6
    ws.row_dimensions[2].height = 15.6
    ws.row_dimensions[3].height = 15.6
    ws.row_dimensions[4].height = 15.6

    ws.cell(1, 1, "Atualização:").font = font_n12
    c = ws.cell(1, 2, atualizacao); c.font = font_b12
    ws.merge_cells("B1:D1")

    ws.cell(2, 1, "Nome:").font = font_n12
    ws.cell(2, 2, nome_banco).font = font_b12

    ws.cell(3, 1, "Agência:").font = font_n12
    ws.cell(3, 2, agencia).font = font_b12

    ws.cell(4, 1, "Conta:").font = font_n12
    c = ws.cell(4, 2, conta); c.font = font_b12
    ws.merge_cells("B4:D4")

    # ── L5: espaçador fino ───────────────────────────────────────────────────
    ws.row_dimensions[5].height = 9.6

    # ── L6: Período ──────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 15.6
    c = ws.cell(6, 1, f"Periodo:  {_periodo_str(fechamento)}")
    c.font = font_n12
    c.fill = fill_periodo

    # ── L7: espaçador muito fino ─────────────────────────────────────────────
    ws.row_dimensions[7].height = 5.4

    # ── L8: cabeçalho da tabela ───────────────────────────────────────────────
    ROW_HEADER = 8
    ws.row_dimensions[ROW_HEADER].height = 37.2
    cabecalho = [
        "DATA", "DESCRIÇÃO LANÇAMENTO BANCO", "DESCRIÇÃO FORNECEDOR/CLIENTE",
        "NF / DOC", "VALOR NF/DOC", "ENTRADA EXTRATO", "SAIDA EXTRATO", "SALDO",
    ]
    for col, label in enumerate(cabecalho, start=1):
        c = ws.cell(ROW_HEADER, col, label)
        c.fill      = fill_azul
        c.font      = font_h
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
    ws.auto_filter.ref = f"A{ROW_HEADER}:H{ROW_HEADER}"

    # ── L9: saldo inicial ─────────────────────────────────────────────────────
    ROW_SALDO = ROW_HEADER + 1
    saldo_ini = _saldo_inicial(lancamentos)
    for col in range(1, 9):
        c = ws.cell(ROW_SALDO, col)
        c.border = borda
        c.font   = font_b11
        c.fill   = fill_saldo
    ws.cell(ROW_SALDO, 1, lancamentos[0].data_lancamento if lancamentos else None)
    ws.cell(ROW_SALDO, 1).number_format = FMT_DATA
    c = ws.cell(ROW_SALDO, 2, "SALDO TOTAL DISPONÍVEL DIA")
    ws.merge_cells(f"B{ROW_SALDO}:G{ROW_SALDO}")
    c = ws.cell(ROW_SALDO, 8, saldo_ini)
    c.number_format = FMT_NUM
    c.alignment     = aln_r
    c.font          = font_b11
    c.fill          = fill_saldo

    # ── L10+: linhas de dados ─────────────────────────────────────────────────
    ROW_INICIO = ROW_SALDO + 1
    for i, l in enumerate(lancamentos):
        row  = ROW_INICIO + i
        prev = row - 1

        entrada = float(l.valor) if l.tipo_movimento == "entrada" else None
        saida   = float(l.valor) if l.tipo_movimento == "saida"   else None
        nf_val  = float(l.valor_nf_doc) if l.valor_nf_doc is not None else None

        # DATA
        c = ws.cell(row, 1, l.data_lancamento)
        c.number_format = FMT_DATA; c.border = borda; c.font = font_n11; c.alignment = aln_c

        # DESCRIÇÃO BANCO
        c = ws.cell(row, 2, l.descricao_banco or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_l

        # DESCRIÇÃO FORNECEDOR/CLIENTE
        c = ws.cell(row, 3, l.descricao_negocio or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_l

        # NF/DOC
        c = ws.cell(row, 4, l.nf_doc or "")
        c.border = borda; c.font = font_n11; c.alignment = aln_c

        # VALOR NF/DOC
        c = ws.cell(row, 5, nf_val)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_n11; c.alignment = aln_r

        # ENTRADA (verde)
        c = ws.cell(row, 6, entrada)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_verde; c.alignment = aln_r

        # SAÍDA (vermelho)
        c = ws.cell(row, 7, saida)
        c.number_format = FMT_NUM; c.border = borda; c.font = font_vermelho; c.alignment = aln_r

        # SALDO (fórmula)
        c = ws.cell(row, 8, f"=H{prev}+F{row}-G{row}")
        c.number_format = FMT_NUM; c.border = borda; c.font = font_n11; c.alignment = aln_r

    # ── Linha final: saldo total disponível ───────────────────────────────────
    last_row  = ROW_INICIO + len(lancamentos) - 1
    row_fim   = last_row + 1
    for col in range(1, 9):
        c = ws.cell(row_fim, col)
        c.border = borda; c.font = font_b11; c.fill = fill_saldo
    ws.cell(row_fim, 1, lancamentos[-1].data_lancamento if lancamentos else None)
    ws.cell(row_fim, 1).number_format = FMT_DATA
    ws.cell(row_fim, 2, "SALDO TOTAL DISPONÍVEL DIA")
    ws.merge_cells(f"B{row_fim}:G{row_fim}")
    c = ws.cell(row_fim, 8, f"=H{last_row}")
    c.number_format = FMT_NUM; c.alignment = aln_r; c.font = font_b11; c.fill = fill_saldo

    # ── Salvar ────────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    conteudo = buffer.read()

    empresa_slug = _sanitizar_nome(empresa_nome)
    data_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    id_curto = str(fechamento.id)[:8]
    nome = f"ia16_extrato_anotado_{empresa_slug}_{data_str}_{id_curto}.xlsx"

    return conteudo, nome


def _gerar_excel_bilateral(db: Session, fechamento: FechamentoFinanceiro) -> tuple[bytes, str]:
    """
    Gera o Excel do pacote final do fechamento.
    Retorna (bytes_do_arquivo, nome_do_arquivo).
    """
    empresa = db.query(Empresa).filter(Empresa.id == fechamento.empresa_id).first()
    empresa_nome = empresa.nome if empresa else str(fechamento.empresa_id)

    aprovado_por: str | None = None
    if fechamento.aprovado_por_usuario_id:
        u = db.query(Usuario).filter(Usuario.id == fechamento.aprovado_por_usuario_id).first()
        aprovado_por = u.email if u else str(fechamento.aprovado_por_usuario_id)

    reaberto_por: str | None = None
    if fechamento.reaberto_por_usuario_id:
        u = db.query(Usuario).filter(Usuario.id == fechamento.reaberto_por_usuario_id).first()
        reaberto_por = u.email if u else str(fechamento.reaberto_por_usuario_id)

    # Buscar itens conciliados
    itens_conciliados = (
        db.query(ItemConciliacao)
        .filter(
            ItemConciliacao.fechamento_id == fechamento.id,
            ItemConciliacao.status == "conciliado",
        )
        .all()
    )

    # Buscar todas as divergências
    todas_divergencias = (
        db.query(DivergenciaConciliacao)
        .filter(DivergenciaConciliacao.fechamento_id == fechamento.id)
        .all()
    )

    divergencias_normais = [d for d in todas_divergencias if d.tipo_divergencia != "pendente_analise_manual"]
    pendencias = [d for d in todas_divergencias if d.tipo_divergencia == "pendente_analise_manual"]
    com_observacao = [d for d in todas_divergencias if d.observacao]

    # Montar workbook
    wb = openpyxl.Workbook()
    ws_resumo = wb.active
    _aba_resumo(ws_resumo, fechamento, empresa_nome, aprovado_por, reaberto_por)

    _aba_itens_conciliados(wb.create_sheet(), itens_conciliados)
    _aba_divergencias(wb.create_sheet(), divergencias_normais)
    _aba_pendencias(wb.create_sheet(), pendencias)
    _aba_observacoes(wb.create_sheet(), com_observacao)

    # Gerar bytes em memória
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    conteudo = buffer.read()

    # Nome do arquivo
    empresa_slug = _sanitizar_nome(empresa_nome)
    data_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    id_curto = str(fechamento.id)[:8]
    nome_arquivo = f"ia16_fechamento_inteligente_{empresa_slug}_{data_str}_{id_curto}.xlsx"

    return conteudo, nome_arquivo
