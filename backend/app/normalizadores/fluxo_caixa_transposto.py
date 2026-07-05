"""
Normalizador para fluxo de caixa em formato transposto diário.

Estrutura esperada:
  - Aba configurável (configurada no modelo, mas detecção é por estrutura)
  - Linha 1: datas nas colunas a partir de D
  - Coluna A: nível da categoria (1, 2, ou vazio)
  - Coluna B: nome da categoria
  - Células: valor previsto para aquela categoria naquela data

Regras:
  - Linhas cujo nome começa com prefixo de totalização (ex: "TOTAL") são ignoradas
  - Linhas de saldo/controle são ignoradas (SALDO INICIAL, ENTRADAS, SAIDAS, etc.)
  - Células None, zero ou string são ignoradas
  - Coluna ACUMULADO (não é date) é ignorada automaticamente
  - Valor positivo = entrada, negativo = saída
  - Células com comentário tentam ser desmembradas em lançamentos individuais
"""
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from app.normalizadores.parser_comentario_celula import parsear_itens_comentario


# Linhas de controle/saldo que não geram lançamentos
_LINHAS_CONTROLE = frozenset({
    'SALDO INICIAL', 'ENTRADAS', 'SAIDAS', 'SAÍDAS',
    'SALDO FINAL', 'TOTAL DISPONIBILIDADES',
})


@dataclass
class LancamentoPrevisto:
    data_prevista: date
    descricao_prevista: str
    categoria: str
    valor_previsto: Decimal
    tipo_movimento: str  # "entrada" | "saida"
    arquivo_id: str
    linha_origem: int
    coluna_origem: str
    metadados: dict[str, Any] = field(default_factory=dict)


def normalizar(conteudo: bytes, config: dict, arquivo_id: str) -> list[LancamentoPrevisto]:
    """
    Normaliza um fluxo de caixa transposto e retorna lista de lançamentos previstos.

    Args:
        conteudo:   Bytes do arquivo .xlsx
        config:     mapeamento_colunas do ModeloArquivo
        arquivo_id: UUID do ArquivoEnviado (como string)
    """
    aba_nome_config    = config.get("aba")
    linha_datas        = config.get("linha_datas", 1)
    col_cat_idx        = column_index_from_string(config.get("coluna_categoria", "B"))
    col_inicio_val_idx = column_index_from_string(config.get("coluna_inicio_valores", "D"))
    prefixos_total     = [p.upper() for p in config.get("prefixos_totalizacao", ["TOTAL"])]

    # data_only=True para obter valores calculados de fórmulas.
    # Não usamos read_only para poder acessar comentários das células.
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)

    ws = _selecionar_aba(wb, aba_nome_config, linha_datas, col_inicio_val_idx)
    if ws is None:
        wb.close()
        raise ValueError(
            f"Nenhuma aba com estrutura de fluxo transposto encontrada. "
            f"Aba esperada: '{aba_nome_config}', abas disponíveis: {wb.sheetnames}"
        )

    # Ler mapa de datas da linha de cabeçalho (col_idx → date)
    datas_por_coluna: dict[int, date] = {}
    for row in ws.iter_rows(min_row=linha_datas, max_row=linha_datas):
        for cell in row:
            col_num = cell.column
            if col_num is None or col_num < col_inicio_val_idx:
                continue
            d = _val_para_date(cell.value)
            if d is not None:
                datas_por_coluna[col_num] = d

    if not datas_por_coluna:
        wb.close()
        raise ValueError(
            f"Nenhuma data encontrada na linha {linha_datas} "
            f"a partir da coluna {get_column_letter(col_inicio_val_idx)}."
        )

    previstos: list[LancamentoPrevisto] = []

    for row in ws.iter_rows(min_row=linha_datas + 1):
        num_linha: int = row[0].row or 0
        if len(row) < col_cat_idx:
            continue

        cat_cell = row[col_cat_idx - 1]
        categoria_raw = cat_cell.value
        if categoria_raw is None:
            continue
        categoria = str(categoria_raw).strip()
        if not categoria:
            continue

        # Ignorar linhas de totalização
        if any(categoria.upper().startswith(p) for p in prefixos_total):
            continue

        # Ignorar linhas de controle/saldo
        if _e_linha_controle(categoria):
            continue

        for cell in row:
            col_idx = cell.column
            if col_idx not in datas_por_coluna:
                continue

            data = datas_por_coluna[col_idx]
            valor_celula = _parse_decimal(cell.value)
            if valor_celula is None or valor_celula == 0:
                continue

            col_letra = get_column_letter(col_idx)
            comentario_texto = cell.comment.text if cell.comment else None

            # Tentar desmembrar via comentário
            if comentario_texto:
                lancamentos = _gerar_de_comentario(
                    comentario_texto=comentario_texto,
                    valor_celula=valor_celula,
                    categoria=categoria,
                    data=data,
                    arquivo_id=arquivo_id,
                    num_linha=num_linha,
                    col_letra=col_letra,
                )
                if lancamentos:
                    previstos.extend(lancamentos)
                    continue

            # Comportamento padrão: lançamento único com valor da célula
            tipo_mov = "entrada" if valor_celula > 0 else "saida"
            meta: dict[str, Any] = {}
            if comentario_texto:
                meta['comentario_original'] = comentario_texto
                meta['celula_origem'] = f"{col_letra}{num_linha}"
                meta['valor_celula'] = str(valor_celula)

            previstos.append(LancamentoPrevisto(
                data_prevista=data,
                descricao_prevista=categoria,
                categoria=categoria,
                valor_previsto=abs(valor_celula),
                tipo_movimento=tipo_mov,
                arquivo_id=arquivo_id,
                linha_origem=num_linha,
                coluna_origem=col_letra,
                metadados=meta,
            ))

    wb.close()
    return previstos


# ── Geração a partir de comentário ────────────────────────────────────────────

def _gerar_de_comentario(
    comentario_texto: str,
    valor_celula: Decimal,
    categoria: str,
    data: date,
    arquivo_id: str,
    num_linha: int,
    col_letra: str,
) -> list[LancamentoPrevisto]:
    """
    Tenta gerar lançamentos individuais a partir do comentário.
    Retorna [] se não conseguir extrair itens confiáveis.
    """
    itens = parsear_itens_comentario(comentario_texto)
    if not itens:
        return []

    soma_itens = sum(i['valor'] for i in itens)
    diferenca = abs(valor_celula) - soma_itens

    lancamentos = []
    for item in itens:
        tipo_mov = "entrada" if valor_celula > 0 else "saida"

        meta: dict[str, Any] = {
            'comentario_original': comentario_texto,
            'celula_origem': f"{col_letra}{num_linha}",
            'valor_celula': str(abs(valor_celula)),
            'item_extraido_de_comentario': True,
            'cliente_ou_descricao': item['descricao'],
            'soma_itens_comentario': str(soma_itens),
            'diferenca_soma_comentario': str(diferenca),
        }
        if item.get('oc'):
            meta['oc'] = item['oc']
        if item.get('pi'):
            meta['pi'] = item['pi']
        if item.get('nf'):
            meta['nf'] = item['nf']
        if item.get('pp'):
            meta['pp'] = item['pp']
        meta['valor_item'] = str(item['valor'])

        desc = item['descricao'] or categoria

        lancamentos.append(LancamentoPrevisto(
            data_prevista=data,
            descricao_prevista=desc,
            categoria=categoria,
            valor_previsto=item['valor'],
            tipo_movimento=tipo_mov,
            arquivo_id=arquivo_id,
            linha_origem=num_linha,
            coluna_origem=col_letra,
            metadados=meta,
        ))

    return lancamentos


# ── Detecção de aba ───────────────────────────────────────────────────────────

def _selecionar_aba(wb, aba_nome: str | None, linha_datas: int, col_inicio_idx: int):
    """Retorna a worksheet correta: por nome se existir, senão pela estrutura."""
    if aba_nome and aba_nome in wb.sheetnames:
        return wb[aba_nome]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        date_count = 0
        for row in ws.iter_rows(min_row=linha_datas, max_row=linha_datas):
            for cell in row:
                if cell.column >= col_inicio_idx and _val_para_date(cell.value) is not None:
                    date_count += 1
        if date_count >= 3:
            return ws
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _val_para_date(val: Any) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    try:
        d = Decimal(str(val)).normalize()
        return d if d != 0 else None
    except (InvalidOperation, ValueError):
        return None


def _e_linha_controle(categoria: str) -> bool:
    upper = categoria.upper()
    return any(ctrl in upper for ctrl in _LINHAS_CONTROLE)
