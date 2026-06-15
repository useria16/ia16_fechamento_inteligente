"""
Normalizador para extratos bancários em formato tabular.

Estrutura esperada:
  - Aba configurável (ex: "Lançamentos")
  - Metadados nas primeiras linhas (nome do banco, conta, etc.)
  - Cabeçalho em linha configurável (ex: linha 10)
  - Dados a partir da linha seguinte ao cabeçalho
  - Colunas: Data | Descrição Operação | Razão Social | Documento | Valor | Saldo

Regras:
  - Linhas de saldo (SALDO ANTERIOR, SALDO TOTAL...) são ignoradas
  - Valor positivo = entrada, negativo = saída
  - Colunas extras além da F são lidas como metadados de debug
"""
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


@dataclass
class LancamentoRealizado:
    data_realizada: date
    descricao_realizada: str
    descricao_operacao: str
    razao_social: str | None
    documento: str | None
    valor_realizado: Decimal
    tipo_movimento: str  # "entrada" | "saida"
    arquivo_id: str
    linha_origem: int
    saldo: Decimal | None = None
    metadados: dict[str, Any] = field(default_factory=dict)


def normalizar(conteudo: bytes, config: dict, arquivo_id: str) -> list[LancamentoRealizado]:
    """
    Normaliza um extrato bancário tabular e retorna lista de lançamentos realizados.

    Args:
        conteudo:   Bytes do arquivo .xlsx
        config:     mapeamento_colunas do ModeloArquivo
        arquivo_id: UUID do ArquivoEnviado (como string)
    """
    aba_nome           = config.get("aba", "Lançamentos")
    linha_cabecalho    = config.get("linha_cabecalho", 10)
    linha_inicio_dados = config.get("linha_inicio_dados", linha_cabecalho + 1)
    col_data_idx       = column_index_from_string(config.get("coluna_data", "A"))
    col_op_idx         = column_index_from_string(config.get("coluna_descricao_operacao", "B"))
    col_razao_idx      = column_index_from_string(config.get("coluna_razao_social", "C"))
    col_doc_idx        = column_index_from_string(config.get("coluna_documento", "D"))
    col_valor_idx      = column_index_from_string(config.get("coluna_valor", "E"))
    col_saldo_str      = config.get("coluna_saldo", "F")
    col_saldo_idx      = column_index_from_string(col_saldo_str) if col_saldo_str else None
    filtros_saldo      = [s.upper() for s in config.get("filtro_linhas_saldo", [
        "SALDO ANTERIOR", "SALDO TOTAL DISPONÍVEL DIA", "SALDO EM CONTA CORRENTE",
    ])]

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)

    if aba_nome not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba '{aba_nome}' não encontrada. Abas disponíveis: {wb.sheetnames}")

    ws = wb[aba_nome]
    realizados: list[LancamentoRealizado] = []

    # Saldo acumulado: inicializado a partir do SALDO ANTERIOR, depois calculado por transação.
    saldo_acumulado: Decimal | None = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        num_linha = i + 1
        if num_linha < linha_inicio_dados:
            continue

        def _cel(idx: int) -> Any:
            return row[idx - 1] if len(row) >= idx else None

        op_val    = _cel(col_op_idx)
        data_val  = _cel(col_data_idx)
        razao_val = _cel(col_razao_idx)
        doc_val   = _cel(col_doc_idx)
        valor_val = _cel(col_valor_idx)
        saldo_val = _cel(col_saldo_idx) if col_saldo_idx else None

        # Pular linhas vazias
        if op_val is None and valor_val is None:
            continue

        op_str_upper = str(op_val or "").strip().upper()

        # Capturar saldo inicial das linhas de saldo antes de filtrá-las
        if op_str_upper in filtros_saldo:
            if saldo_val is not None and saldo_acumulado is None:
                saldo_inicial = _parse_decimal(saldo_val)
                if saldo_inicial is not None:
                    saldo_acumulado = saldo_inicial
            continue

        if valor_val is None:
            continue

        # Parse data
        data = _parse_data(data_val)
        if data is None:
            continue

        # Parse valor
        valor = _parse_decimal(valor_val)
        if valor is None or valor == 0:
            continue

        tipo_mov = "entrada" if valor > 0 else "saida"

        # Calcular saldo acumulado após este lançamento
        if saldo_acumulado is not None:
            saldo_acumulado = saldo_acumulado + valor
        elif saldo_val is not None:
            # Fallback: coluna de saldo preenchida diretamente no lançamento
            saldo_acumulado = _parse_decimal(saldo_val)

        # Descrição combinada: operação + razão social
        op_str    = str(op_val or "").strip()
        razao_str = str(razao_val or "").strip() if razao_val else None
        descricao = f"{op_str} {razao_str}".strip() if razao_str else op_str

        # Metadados: colunas extras após a coluna de saldo
        inicio_extras = (col_saldo_idx or col_valor_idx)
        metadados: dict[str, Any] = {}
        for extra_idx in range(inicio_extras, len(row)):
            extra_val = row[extra_idx]
            if extra_val is not None:
                metadados[f"col_{get_column_letter(extra_idx + 1)}"] = str(extra_val)

        realizados.append(LancamentoRealizado(
            data_realizada=data,
            descricao_realizada=descricao,
            descricao_operacao=op_str,
            razao_social=razao_str if razao_str else None,
            documento=str(doc_val).strip() if doc_val else None,
            valor_realizado=valor,
            tipo_movimento=tipo_mov,
            arquivo_id=arquivo_id,
            linha_origem=num_linha,
            saldo=saldo_acumulado,
            metadados=metadados,
        ))

    wb.close()
    return realizados


def _parse_data(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val)).normalize()
    except (InvalidOperation, ValueError):
        return None
