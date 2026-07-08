"""
Testes do normalizador de fluxo de caixa transposto e do parser de comentários.

Cobre:
  - Parser: sem comentário, item único, múltiplos itens, datas no comentário
  - Normalização: célula sem comentário, célula com comentário, linhas de saldo
  - Diferença entre soma dos itens e valor da célula
  - Cópia temporária da planilha real (sem alterar o original)
"""
import io
import os
import shutil
import tempfile
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from openpyxl.comments import Comment

from app.normalizadores.parser_comentario_celula import parsear_itens_comentario
from app.normalizadores.fluxo_caixa_transposto import normalizar, LancamentoPrevisto


# ── Fixtures de config ────────────────────────────────────────────────────────

CONFIG_PADRAO = {
    "aba": "Fluxo",
    "linha_datas": 1,
    "coluna_categoria": "B",
    "coluna_inicio_valores": "D",
    "prefixos_totalizacao": ["TOTAL"],
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _wb_em_memoria(linhas: list[list], comentarios: dict | None = None) -> bytes:
    """
    Cria workbook em memória.
    linhas: lista de linhas, cada linha é lista de valores (coluna B em diante via offset).
    comentarios: dict {(row, col): "texto"} com coordenadas 1-based.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fluxo"

    for r_idx, linha in enumerate(linhas, 1):
        for c_idx, val in enumerate(linha, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    if comentarios:
        for (row, col), texto in comentarios.items():
            cell = ws.cell(row=row, column=col)
            cell.comment = Comment(texto, "sistema")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Testes do parser de comentários ──────────────────────────────────────────

class TestParserComentario:

    def test_texto_vazio_retorna_lista_vazia(self):
        assert parsear_itens_comentario("") == []
        assert parsear_itens_comentario("   ") == []

    def test_sem_valor_retorna_lista_vazia(self):
        texto = "Autor\nTexto sem valor monetário"
        assert parsear_itens_comentario(texto) == []

    def test_item_unico_com_rs(self):
        texto = "Alice Santos\nABA OC 696/A R$ 5.525,00 NF 235 TOTAL R$ 5.525,00"
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 1
        assert itens[0]['valor'] == Decimal('5525.00')
        assert itens[0]['oc'] == '696/A'
        assert itens[0]['nf'] == '235'

    def test_multiplos_itens_linhas_separadas(self):
        texto = (
            "Alice Santos\n"
            "ESTRATEGIA CONCURSOS PI 6939 ME R$ 2.498,05\n"
            "ESTRATEGIA CONCURSOS PI 6940 ME R$ 1.386,96\n"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 2
        assert itens[0]['valor'] == Decimal('2498.05')
        assert itens[0]['pi'] == '6939'
        assert itens[1]['valor'] == Decimal('1386.96')
        assert itens[1]['pi'] == '6940'

    def test_multiplos_itens_mesmo_bloco(self):
        texto = (
            "Alice Santos\n"
            "ABA OC 696/A R$ 5.525,00 NF 235 TOTAL R$ 5.525,00\n"
            "\n"
            "ABA OC 711/A R$ 2.975,00 NF 236 TOTAL R$ 2.975,00\n"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 2
        valores = {i['valor'] for i in itens}
        assert Decimal('5525.00') in valores
        assert Decimal('2975.00') in valores

    def test_formato_sinalizado_multiplos(self):
        texto = (
            "Cristina\n"
            "+ 4.000 - João Pedro - copy e edição - PP 1013 - 3 de 4 "
            "+ 2.000 - Lucas - designer - PP 1014 - 3 DE 4"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 2
        valores = {i['valor'] for i in itens}
        assert Decimal('4000') in valores
        assert Decimal('2000') in valores

    def test_datas_nao_viram_valores(self):
        texto = (
            "Alice Santos - atualizado 08/01/26\n"
            "ABA PI 6862 ME 2.119,91 NF 222 TOTAL R$ 2.119,91\n"
            "******* 05/01/26 - algum texto 03/12/25\n"
        )
        itens = parsear_itens_comentario(texto)
        # Só deve extrair o item com TOTAL R$
        assert len(itens) == 1
        assert itens[0]['valor'] == Decimal('2119.91')

    def test_linhas_controle_ignoradas(self):
        texto = (
            "Sistema\n"
            "SALDO INICIAL R$ 10.000,00\n"
            "ENTRADAS R$ 5.000,00\n"
            "ABA OC 100 R$ 3.000,00 TOTAL R$ 3.000,00\n"
            "SALDO FINAL R$ 8.000,00\n"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 1
        assert itens[0]['valor'] == Decimal('3000.00')

    def test_item_com_total_usa_valor_total(self):
        # Garante que TOTAL R$ tem precedência sobre primeiro R$
        texto = (
            "Autor\n"
            "DI GASPI OC 1012 NF 246 TOTAL R$ 17.000,00\n"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 1
        assert itens[0]['valor'] == Decimal('17000.00')
        assert itens[0]['oc'] == '1012'
        assert itens[0]['nf'] == '246'

    def test_quatro_itens_estrategia(self):
        texto = (
            "Alice Santos - atualizado 08/01/26\n"
            "\n"
            "ESTRATEGIA CONCURSOS PI 6939 ME  R$ 2.498,05 \n"
            "ESTRATEGIA CONCURSOS PI 6940 ME  R$ 1.386,96 \n"
            "ESTRATEGIA CONCURSOS PI 6941 ME  R$ 2.381,66 \n"
            "ESTRATEGIA CONCURSOS PI 6942 ME  R$ 114,76 NF 234   TOTAL R$ 6.381,43\n"
        )
        itens = parsear_itens_comentario(texto)
        assert len(itens) == 4
        soma = sum(i['valor'] for i in itens)
        assert [i['valor'] for i in itens] == [
            Decimal('2498.05'),
            Decimal('1386.96'),
            Decimal('2381.66'),
            Decimal('114.76'),
        ]
        assert soma == Decimal('6381.43')


# ── Testes de normalização com workbook em memória ────────────────────────────

class TestNormalizadorFluxoCaixa:

    def _conteudo_basico(self, comentarios=None):
        """
        Monta workbook:
        L1: B=cabeçalho  D=2026-01-01  E=2026-01-02
        L2: B=SALDO INICIAL  D=1000  E=2000   (linha controle, ignorar)
        L3: B=Receita Op.    D=5000  E=None
        L4: B=TOTAL RECEITAS D=6000  E=2000   (prefixo TOTAL, ignorar)
        L5: B=Custo Variável D=-3000 E=-1500
        """
        from datetime import datetime as dt
        linhas = [
            [None, "Fluxo Jan",    None,    dt(2026, 1, 1), dt(2026, 1, 2)],
            [None, "SALDO INICIAL",None,    1000,           2000           ],
            [None, "Receita Op.",  None,    5000,           None           ],
            [None, "TOTAL RECEITAS",None,   6000,           2000           ],
            [None, "Custo Variável",None,  -3000,          -1500           ],
        ]
        return _wb_em_memoria(linhas, comentarios)

    def test_sem_comentarios(self):
        conteudo = self._conteudo_basico()
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-001")
        # Receita Op. D (5000) + Custo Variável D (-3000) + Custo Variável E (-1500)
        assert len(resultado) == 3
        valores = {abs(r.valor_previsto) for r in resultado}
        assert Decimal('5000') in valores
        assert Decimal('3000') in valores
        assert Decimal('1500') in valores

    def test_linhas_saldo_ignoradas(self):
        conteudo = self._conteudo_basico()
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-001")
        categorias = {r.categoria for r in resultado}
        assert 'SALDO INICIAL' not in categorias
        assert 'TOTAL RECEITAS' not in categorias

    def test_tipo_movimento(self):
        conteudo = self._conteudo_basico()
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-001")
        entradas = [r for r in resultado if r.tipo_movimento == 'entrada']
        saidas   = [r for r in resultado if r.tipo_movimento == 'saida']
        assert len(entradas) == 1
        assert len(saidas) == 2

    def test_celula_com_comentario_item_unico(self):
        from datetime import datetime as dt
        comentario = "Alice\nABA OC 696/A R$ 5.525,00 NF 235 TOTAL R$ 5.525,00"
        linhas = [
            [None, "Header",     None, dt(2026, 1, 1)],
            [None, "Fornecedor", None, 5525           ],
        ]
        # D2 = row=2, col=4
        conteudo = _wb_em_memoria(linhas, {(2, 4): comentario})
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-002")
        assert len(resultado) == 1
        assert resultado[0].valor_previsto == Decimal('5525.00')
        assert resultado[0].metadados.get('oc') == '696/A'
        assert resultado[0].metadados.get('nf') == '235'
        assert resultado[0].metadados['item_extraido_de_comentario'] is True

    def test_celula_com_comentario_multiplos_itens(self):
        from datetime import datetime as dt
        comentario = (
            "Alice\n"
            "ESTRATEGIA CONCURSOS PI 6939 ME R$ 2.498,05\n"
            "ESTRATEGIA CONCURSOS PI 6940 ME R$ 1.386,96\n"
        )
        linhas = [
            [None, "Header",     None, dt(2026, 2, 1)],
            [None, "Fornecedor", None, 3885.01        ],
        ]
        conteudo = _wb_em_memoria(linhas, {(2, 4): comentario})
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-003")
        assert len(resultado) == 2
        soma = sum(r.valor_previsto for r in resultado)
        assert soma == Decimal('2498.05') + Decimal('1386.96')

    def test_celula_sem_comentario_nao_quebra(self):
        from datetime import datetime as dt
        linhas = [
            [None, "Header",     None, dt(2026, 1, 1)],
            [None, "Receita Op.",None, 10000          ],
        ]
        conteudo = _wb_em_memoria(linhas)
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-004")
        assert len(resultado) == 1
        assert resultado[0].valor_previsto == Decimal('10000')
        assert resultado[0].metadados == {}

    def test_diferenca_soma_registrada_em_metadados(self):
        from datetime import datetime as dt
        # Valor da célula: 10.000 | Itens no comentário somam: 9.999,99
        comentario = (
            "Autor\n"
            "ITEM A R$ 5.000,00\n"
            "ITEM B R$ 4.999,99\n"
        )
        linhas = [
            [None, "Header",     None, dt(2026, 1, 1)],
            [None, "Fornecedor", None, 10000          ],
        ]
        conteudo = _wb_em_memoria(linhas, {(2, 4): comentario})
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-005")
        assert len(resultado) == 2
        # Diferença registrada em ambos os lançamentos
        for r in resultado:
            diferenca = Decimal(r.metadados['diferenca_soma_comentario'])
            assert diferenca != Decimal('0')

    def test_comentario_sem_valor_gera_lancamento_unico(self):
        from datetime import datetime as dt
        comentario = "Alice\nTexto livre sem valor monetário algum"
        linhas = [
            [None, "Header",     None, dt(2026, 1, 1)],
            [None, "Fornecedor", None, 7500           ],
        ]
        conteudo = _wb_em_memoria(linhas, {(2, 4): comentario})
        resultado = normalizar(conteudo, CONFIG_PADRAO, "arq-006")
        # Sem itens extraídos → lançamento único com comentário em metadados
        assert len(resultado) == 1
        assert resultado[0].valor_previsto == Decimal('7500')
        assert 'comentario_original' in resultado[0].metadados


# ── Teste com cópia temporária da planilha real ───────────────────────────────

PLANILHA_REAL = os.getenv("IA16_PLANILHA_REAL_FLUXO_CAIXA", "")

CONFIG_REAL = {
    "aba": "DAXX SERVIÇOS",
    "linha_datas": 1,
    "coluna_categoria": "B",
    "coluna_inicio_valores": "D",
    "prefixos_totalizacao": ["TOTAL"],
}


@pytest.mark.skipif(
    not PLANILHA_REAL or not os.path.exists(PLANILHA_REAL),
    reason="Defina IA16_PLANILHA_REAL_FLUXO_CAIXA para testar com a planilha real",
)
class TestPlanilhaReal:

    @pytest.fixture(autouse=True)
    def copia_temporaria(self, tmp_path):
        """Cria cópia temporária — nunca altera o original."""
        destino = tmp_path / "fluxo_caixa_copia.xlsx"
        shutil.copy2(PLANILHA_REAL, destino)
        self.caminho_copia = str(destino)

    def _conteudo(self) -> bytes:
        with open(self.caminho_copia, 'rb') as f:
            return f.read()

    def test_normaliza_sem_erros(self):
        resultado = normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        assert isinstance(resultado, list)
        assert len(resultado) > 0

    def test_nenhum_lancamento_com_categoria_saldo(self):
        resultado = normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        categorias_proibidas = {'SALDO INICIAL', 'ENTRADAS', 'SAIDAS', 'SAÍDAS',
                                'SALDO FINAL', 'TOTAL DISPONIBILIDADES'}
        for r in resultado:
            assert r.categoria.upper() not in categorias_proibidas

    def test_celulas_com_comentario_geram_itens_detalhados(self):
        resultado = normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        itens_detalhados = [r for r in resultado if r.metadados.get('item_extraido_de_comentario')]
        assert len(itens_detalhados) > 0

    def test_original_nao_modificado(self):
        import hashlib
        with open(PLANILHA_REAL, 'rb') as f:
            hash_antes = hashlib.md5(f.read()).hexdigest()
        normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        with open(PLANILHA_REAL, 'rb') as f:
            hash_depois = hashlib.md5(f.read()).hexdigest()
        assert hash_antes == hash_depois, "Planilha original foi modificada!"

    def test_todos_valores_positivos(self):
        resultado = normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        for r in resultado:
            assert r.valor_previsto > 0, f"Valor negativo: {r}"

    def test_tipos_movimento_validos(self):
        resultado = normalizar(self._conteudo(), CONFIG_REAL, "arq-real-001")
        for r in resultado:
            assert r.tipo_movimento in ('entrada', 'saida')
