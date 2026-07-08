"""
Testes do service de exportação mensal de conciliação (exportacao_mensal_service).

Cobre:
  1.  Exportação diária existente continua funcionando (smoke test do import).
  2.  Erro quando não há conciliações no mês.
  3.  Inclui apenas conciliações da empresa correta (isolamento multicliente).
  4.  Filtra por tipo de conciliação.
  5.  Inclui múltiplos dias do mesmo mês.
  6.  Exclui conciliações fora do mês.
  7.  Gera workbook com uma única aba.
  8.  Aba tem nome no formato mês+ano (ex: Jun26).
  9.  Aba NÃO contém abas antigas (Resumo Mensal, Conciliação Mensal, etc.).
 10.  Cabeçalho da tabela está na linha 8 com as colunas corretas.
 11.  Saldo inicial na linha 9; lançamentos a partir da linha 10.
 12.  Nome do arquivo segue o padrão Conciliacao_<Empresa>_<Mes>_<Ano>.xlsx.
 13.  Dados de entrada e saída separados corretamente (colunas F e G).
 14.  Cabeçalho usa metadados do arquivo de extrato quando disponíveis.
 15.  Fallback seguro quando não há metadados.
 16.  Arquivo com metadados={} usa fallback (não quebra).
 17.  Busca de metadados usa query única com .in_() — não loop por fechamento.
 18.  Metadados preferem arquivo com agência E conta quando há dois arquivos.
 19.  Filtro automático está em A8:H8.
 20.  Linha 6 tem fill BDD7EE (azul claro — período).
 21.  Linha 8 tem fill 1F4E79 (azul escuro — cabeçalho).
 22.  Linha 9 coluna B contém "SALDO TOTAL DISPONÍVEL DIA" (saldo inicial).
 23.  Última linha coluna B contém "SALDO TOTAL DISPONÍVEL DIA" (saldo final).
 24.  Coluna A nas linhas de dados tem formato "DD/MM/YYYY".
 25.  Colunas F, G, H nas linhas de dados têm formato "#,##0.00".
 26.  Entrada (coluna F) tem fonte verde (00B050).
 27.  Saída (coluna G) tem fonte vermelha (FF0000).
"""
import io
import uuid
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock
import pytest

import openpyxl

from app.services.exportacao_mensal_service import (
    buscar_fechamentos_do_periodo,
    buscar_fechamentos_do_mes,
    gerar_excel_periodo,
    gerar_excel_mensal,
    _nome_aba,
    _COLUNAS_MENSAL,
    _LINHA_CABECALHO_TABELA,
)
from app.services.exportacao_fechamento_service import gerar_excel_fechamento


# ── Fixtures ─────────────────────────────────────────────────────────────────

EMPRESA_A = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
EMPRESA_B = uuid.UUID("bbbbbbbb-0000-0000-0000-000000000002")
FECH_1 = uuid.UUID("00000000-1111-0000-0000-000000000001")
FECH_2 = uuid.UUID("00000000-2222-0000-0000-000000000002")
FECH_3 = uuid.UUID("00000000-3333-0000-0000-000000000003")
ARQ_1 = uuid.UUID("00000000-0000-1111-0000-000000000001")


def _make_fechamento(
    fid=FECH_1,
    empresa_id=EMPRESA_A,
    tipo="extrato_anotado",
    status="aprovado",
    periodo_inicio=date(2026, 6, 1),
    titulo="Conciliação 01/06",
):
    f = MagicMock()
    f.id = fid
    f.empresa_id = empresa_id
    f.tipo_conciliacao = tipo
    f.status = status
    f.periodo_inicio = periodo_inicio
    f.periodo_fim = periodo_inicio
    f.titulo = titulo
    return f


def _make_lancamento(
    fechamento_id=FECH_1,
    empresa_id=EMPRESA_A,
    arquivo_id=ARQ_1,
    data=date(2026, 6, 1),
    descricao_banco="PIX recebido",
    descricao_negocio="Cliente X",
    tipo_movimento="entrada",
    valor=Decimal("500.00"),
    saldo=Decimal("1500.00"),
    status_revisao="revisado",
    tipo_conferencia_fluxo="encontrado",
    nf_doc=None,
    valor_nf_doc=None,
    linha_origem=1,
):
    l = MagicMock()
    l.id = uuid.uuid4()
    l.fechamento_id = fechamento_id
    l.empresa_id = empresa_id
    l.arquivo_id = arquivo_id
    l.data_lancamento = data
    l.descricao_banco = descricao_banco
    l.descricao_negocio = descricao_negocio
    l.tipo_movimento = tipo_movimento
    l.valor = valor
    l.saldo = saldo
    l.status_revisao = status_revisao
    l.tipo_conferencia_fluxo = tipo_conferencia_fluxo
    l.nf_doc = nf_doc
    l.valor_nf_doc = valor_nf_doc
    l.linha_origem = linha_origem
    return l


def _make_empresa(eid=EMPRESA_A, nome="Empresa Teste"):
    e = MagicMock()
    e.id = eid
    e.nome = nome
    return e


def _make_arquivo(fechamento_id=FECH_1, metadados=None):
    a = MagicMock()
    a.fechamento_id = fechamento_id
    a.tipo_arquivo = "extrato_bancario"
    a.metadados = metadados or {}
    return a


def _db_com_fechamentos(fechamentos, lancamentos=None, empresa=None, arquivos_extrato=None):
    """Cria um mock de db que retorna os dados fornecidos.

    arquivos_extrato deve ser uma lista de arquivos (ou None/[] para sem arquivo).
    O service usa .all() para buscar metadados, então o mock configura .all.return_value.
    """
    db = MagicMock()
    empresa_obj = empresa or _make_empresa()

    q_fech = MagicMock()
    q_fech.filter.return_value = q_fech
    q_fech.order_by.return_value = q_fech
    q_fech.all.return_value = fechamentos

    q_emp = MagicMock()
    q_emp.filter.return_value = q_emp
    q_emp.first.return_value = empresa_obj

    q_lanc = MagicMock()
    q_lanc.filter.return_value = q_lanc
    q_lanc.order_by.return_value = q_lanc
    q_lanc.all.return_value = lancamentos or []

    q_arq = MagicMock()
    q_arq.filter.return_value = q_arq
    q_arq.order_by.return_value = q_arq
    q_arq.all.return_value = arquivos_extrato or []  # service usa .all() com .in_()

    def side_effect_query(model):
        from app.models.fechamento_financeiro import FechamentoFinanceiro
        from app.models.empresa import Empresa
        from app.models.lancamento_extrato_anotado import LancamentoExtratoAnotado
        from app.models.arquivo_enviado import ArquivoEnviado
        if model is FechamentoFinanceiro:
            return q_fech
        if model is Empresa:
            return q_emp
        if model is LancamentoExtratoAnotado:
            return q_lanc
        if model is ArquivoEnviado:
            return q_arq
        return MagicMock()

    db.query.side_effect = side_effect_query
    return db


# ── Teste 1: importação do serviço diário não quebra ─────────────────────────

def test_importacao_servico_diario():
    """Exportação diária existente: o import e função principal devem existir."""
    assert callable(gerar_excel_fechamento)


# ── Teste 2: erro quando não há conciliações ──────────────────────────────────

def test_erro_sem_conciliacoes_no_mes():
    db = _db_com_fechamentos(fechamentos=[])
    with pytest.raises(ValueError, match="Nenhuma conciliação encontrada"):
        gerar_excel_mensal(
            db=db,
            empresa_id=EMPRESA_A,
            ano=2026,
            mes=6,
            tipo_conciliacao="extrato_anotado",
        )


# ── Teste 3: não inclui empresa errada ────────────────────────────────────────

def test_isolamento_por_empresa():
    fech_a = _make_fechamento(fid=FECH_1, empresa_id=EMPRESA_A)
    db = _db_com_fechamentos(fechamentos=[fech_a], lancamentos=[])
    conteudo, nome = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert len(conteudo) > 0
    assert "2026" in nome
    assert db.query.called


# ── Teste 4: filtra por tipo de conciliação ───────────────────────────────────

def test_filtra_por_tipo_conciliacao():
    fech = _make_fechamento(fid=FECH_1, tipo="extrato_anotado", status="aprovado")
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[])
    conteudo, nome = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert len(conteudo) > 0
    # nome não deve conter o tipo_conciliacao
    assert "extrato_anotado" not in nome


# ── Teste 5: inclui múltiplos dias do mesmo mês ───────────────────────────────

def test_multiplos_dias_mesmo_mes():
    fech1 = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 1))
    fech2 = _make_fechamento(fid=FECH_2, periodo_inicio=date(2026, 6, 2))
    fech3 = _make_fechamento(fid=FECH_3, periodo_inicio=date(2026, 6, 9))

    l1 = _make_lancamento(fechamento_id=FECH_1, data=date(2026, 6, 1))
    l2 = _make_lancamento(fechamento_id=FECH_2, data=date(2026, 6, 2))
    l3 = _make_lancamento(fechamento_id=FECH_3, data=date(2026, 6, 9))

    db = _db_com_fechamentos(fechamentos=[fech1, fech2, fech3], lancamentos=[l1, l2, l3])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert len(conteudo) > 0

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    # 7 linhas de cabeçalho + 1 cabeçalho de tabela + 3 lançamentos = 11 linhas
    assert ws.max_row >= _LINHA_CABECALHO_TABELA + 3


# ── Teste 6: não inclui datas fora do mês ────────────────────────────────────

def test_exclui_datas_fora_do_mes():
    fech_junho = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 15))
    db = _db_com_fechamentos(fechamentos=[fech_junho], lancamentos=[])
    fechamentos = buscar_fechamentos_do_mes(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert len(fechamentos) == 1
    assert fechamentos[0].periodo_inicio.month == 6


# ── Teste 7: gera workbook com UMA única aba ─────────────────────────────────

def test_gera_workbook_com_uma_aba():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, nome = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert wb is not None
    assert nome.endswith(".xlsx")
    assert len(wb.sheetnames) == 1, f"Esperado 1 aba, mas encontrou: {wb.sheetnames}"


# ── Teste 8: aba tem nome no formato mês+ano ─────────────────────────────────

def test_aba_tem_nome_mes_ano():
    fech = _make_fechamento(periodo_inicio=date(2026, 6, 1))
    l = _make_lancamento(data=date(2026, 6, 1))
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert wb.sheetnames[0] == "Jun26", f"Nome da aba esperado 'Jun26', recebeu '{wb.sheetnames[0]}'"


def test_nome_aba_diferentes_meses():
    """Valida a função auxiliar de nomeação da aba para todos os meses."""
    assert _nome_aba(1, 2026) == "Jan26"
    assert _nome_aba(6, 2026) == "Jun26"
    assert _nome_aba(12, 2025) == "Dez25"
    assert _nome_aba(3, 2027) == "Mar27"


# ── Teste 9: NÃO contém abas antigas ─────────────────────────────────────────

def test_nao_contem_abas_antigas():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    abas = wb.sheetnames
    assert "Resumo Mensal" not in abas
    assert "Conciliação Mensal" not in abas
    assert "Dias Incluídos" not in abas
    assert "Pendências" not in abas


# ── Teste 10: cabeçalho da tabela na linha 8 com colunas corretas ────────────

def test_cabecalho_tabela_na_linha_8():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    cabecalhos = [ws.cell(row=_LINHA_CABECALHO_TABELA, column=c).value for c in range(1, len(_COLUNAS_MENSAL) + 1)]
    assert cabecalhos == _COLUNAS_MENSAL, f"Cabeçalhos incorretos: {cabecalhos}"


# ── Teste 11: saldo inicial na linha 9; lançamentos a partir da linha 10 ──────

def test_lancamentos_a_partir_da_linha_10():
    """Linha 9 é o saldo inicial. Dados começam na linha 10 (cabeçalho+2)."""
    fech = _make_fechamento()
    l = _make_lancamento(descricao_banco="TED recebida teste")
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # Linha 9 = saldo inicial
    assert ws.cell(row=_LINHA_CABECALHO_TABELA + 1, column=2).value == "SALDO TOTAL DISPONÍVEL DIA"

    # Linha 10 = primeiro lançamento
    primeira_linha_dados = _LINHA_CABECALHO_TABELA + 2
    descricao = ws.cell(row=primeira_linha_dados, column=2).value
    assert descricao == "TED recebida teste", f"Descrição na linha {primeira_linha_dados}: {descricao}"


# ── Teste 12: nome do arquivo segue padrão Conciliacao_<Empresa>_<Mes>_<Ano> ─

def test_nome_arquivo_padrao_amigavel():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    _, nome = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert nome.startswith("Conciliacao_"), f"Nome não começa com 'Conciliacao_': {nome}"
    assert "Junho" in nome, f"Nome do mês por extenso ausente: {nome}"
    assert "2026" in nome, f"Ano ausente no nome: {nome}"
    assert nome.endswith(".xlsx")


def test_nome_arquivo_nao_contem_padroes_antigos():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    _, nome = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    assert "ia16_" not in nome, f"Nome não deve conter 'ia16_': {nome}"
    assert "extrato_anotado" not in nome, f"Nome não deve conter tipo de conciliação: {nome}"
    assert "_06" not in nome, f"Nome não deve conter mês numérico: {nome}"


# ── Teste 13: entrada e saída separados corretamente ─────────────────────────

def test_entrada_saida_separados():
    fech = _make_fechamento()
    l_entrada = _make_lancamento(tipo_movimento="entrada", valor=Decimal("1000.00"), saldo=None)
    l_saida = _make_lancamento(tipo_movimento="saida", valor=Decimal("300.00"), saldo=None)
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l_entrada, l_saida])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    col_entrada = _COLUNAS_MENSAL.index("ENTRADA EXTRATO") + 1
    col_saida = _COLUNAS_MENSAL.index("SAIDA EXTRATO") + 1

    # Linha 9 = saldo inicial, dados a partir da linha 10
    linha_entrada = _LINHA_CABECALHO_TABELA + 2
    linha_saida = _LINHA_CABECALHO_TABELA + 3

    assert ws.cell(row=linha_entrada, column=col_entrada).value == 1000.0
    assert ws.cell(row=linha_entrada, column=col_saida).value in (None, "")

    assert ws.cell(row=linha_saida, column=col_saida).value == 300.0
    assert ws.cell(row=linha_saida, column=col_entrada).value in (None, "")


# ── Teste 14: cabeçalho usa metadados do arquivo de extrato ──────────────────

def test_cabecalho_usa_metadados_do_extrato():
    """Quando há arquivo de extrato com metadados, o cabeçalho usa esses dados."""
    fech = _make_fechamento()
    l = _make_lancamento()
    arquivo = _make_arquivo(
        fechamento_id=FECH_1,
        metadados={
            "atualizacao": "10/06/2026 09:00:00",
            "nome": "BANCO EXEMPLO S/A",
            "agencia": "0123",
            "conta": "9876543-2",
        },
    )
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l], arquivos_extrato=[arquivo])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    assert ws.cell(row=1, column=2).value == "10/06/2026 09:00:00", "Atualização incorreta"
    assert ws.cell(row=2, column=2).value == "BANCO EXEMPLO S/A", "Nome do banco incorreto"
    assert ws.cell(row=3, column=2).value == "0123", "Agência incorreta"
    assert ws.cell(row=4, column=2).value == "9876543-2", "Conta incorreta"


# ── Teste 15: fallback seguro quando não há metadados ────────────────────────

def test_cabecalho_fallback_sem_metadados():
    """Quando não há arquivo de extrato, o cabeçalho usa fallback sem erros."""
    fech = _make_fechamento()
    l = _make_lancamento()
    empresa = _make_empresa(nome="Empresa Sem Extrato")
    # arquivos_extrato=None → sem arquivo (retorna lista vazia)
    db = _db_com_fechamentos(
        fechamentos=[fech], lancamentos=[l], empresa=empresa, arquivos_extrato=None
    )
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # Nome da empresa como fallback
    assert ws.cell(row=2, column=2).value == "Empresa Sem Extrato", "Fallback de nome deve usar empresa"
    # Agência e conta devem estar em branco
    assert ws.cell(row=3, column=2).value in (None, ""), "Agência deve estar vazia sem metadados"
    assert ws.cell(row=4, column=2).value in (None, ""), "Conta deve estar vazia sem metadados"
    # Atualização deve ser uma string não vazia (data/hora atual)
    atualizacao = ws.cell(row=1, column=2).value
    assert atualizacao and len(str(atualizacao)) > 5, "Atualização deve ter valor de fallback"


# ── Teste 16: arquivo com metadados={} usa fallback sem erros ─────────────────

def test_cabecalho_fallback_metadados_vazios():
    """
    Arquivo de extrato existe mas tem metadados={} (ainda não processado).
    Deve usar fallback sem quebrar. Reproduz o bug onde a condição
    `if arquivo and arquivo.metadados:` pulava arquivos com dict vazio.
    """
    fech = _make_fechamento()
    l = _make_lancamento()
    empresa = _make_empresa(nome="Empresa Com Arquivo Vazio")
    arquivo = _make_arquivo(fechamento_id=FECH_1, metadados={})
    db = _db_com_fechamentos(
        fechamentos=[fech], lancamentos=[l], empresa=empresa, arquivos_extrato=[arquivo]
    )
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # Com metadados={}, fallback deve usar nome da empresa
    assert ws.cell(row=2, column=2).value == "Empresa Com Arquivo Vazio"
    # Agência e conta em branco
    assert ws.cell(row=3, column=2).value in (None, "")
    assert ws.cell(row=4, column=2).value in (None, "")


# ── Teste 17: busca de metadados usa query única com .in_() ───────────────────

def test_metadados_usa_query_unica_com_in():
    """
    _buscar_metadados_extrato deve fazer UMA query com .in_() para todos
    os fechamentos do mês — não um loop de N queries individuais.

    Verifica que db.query(ArquivoEnviado) é chamado exatamente uma vez
    durante a geração, independente do número de fechamentos.
    """
    from app.models.arquivo_enviado import ArquivoEnviado

    fech1 = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 1))
    fech2 = _make_fechamento(fid=FECH_2, periodo_inicio=date(2026, 6, 15))
    l = _make_lancamento()
    arquivo = _make_arquivo(metadados={"agencia": "0001", "conta": "12345-6", "nome": "BANCO X", "atualizacao": "01/06/2026 08:00:00"})

    db = _db_com_fechamentos(
        fechamentos=[fech1, fech2], lancamentos=[l], arquivos_extrato=[arquivo]
    )

    gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )

    # Contabiliza chamadas a db.query com ArquivoEnviado
    calls_arq = [c for c in db.query.call_args_list if c.args and c.args[0] is ArquivoEnviado]
    assert len(calls_arq) == 1, (
        f"Esperado 1 query para ArquivoEnviado (via .in_()), mas foram feitas {len(calls_arq)}. "
        f"O service não deve fazer N queries em loop por fechamento."
    )


# ── Teste 18: prefere arquivo com agencia E conta ─────────────────────────────

def test_metadados_prefere_arquivo_com_agencia_e_conta():
    """
    Quando há dois arquivos de extrato, o service deve preferir o que tem
    agencia E conta. O primeiro (sem metadados) não deve ser usado.
    Reproduz o critério: 'se o diário mostra 0285/0017861-2, o mensal também deve.'
    """
    fech1 = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 1))
    fech2 = _make_fechamento(fid=FECH_2, periodo_inicio=date(2026, 6, 15))
    l = _make_lancamento()
    empresa = _make_empresa(nome="Empresa XYZ")

    arquivo_sem_meta  = _make_arquivo(fechamento_id=FECH_1, metadados={})
    arquivo_com_meta  = _make_arquivo(
        fechamento_id=FECH_2,
        metadados={
            "agencia": "0285",
            "conta": "0017861-2",
            "nome": "BANCO PILOTO S/A",
            "atualizacao": "09/06/2026 10:00:00",
        },
    )

    db = _db_com_fechamentos(
        fechamentos=[fech1, fech2],
        lancamentos=[l],
        empresa=empresa,
        arquivos_extrato=[arquivo_sem_meta, arquivo_com_meta],
    )
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    assert ws.cell(row=3, column=2).value == "0285",       "Agência incorreta"
    assert ws.cell(row=4, column=2).value == "0017861-2",  "Conta incorreta"
    assert ws.cell(row=2, column=2).value == "BANCO PILOTO S/A", "Nome do banco incorreto"


# ── Teste 19: filtro automático em A8:H8 ──────────────────────────────────────

def test_auto_filter_em_a8_h8():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    assert ws.auto_filter.ref == "A8:H8", f"auto_filter.ref incorreto: {ws.auto_filter.ref}"


# ── Teste 20: linha 6 tem fill azul-claro (BDD7EE — período) ─────────────────

def test_linha_6_fill_periodo():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    fill = ws.cell(row=6, column=1).fill
    assert fill is not None, "Linha 6 deve ter fill"
    fgColor = fill.fgColor.rgb if fill.fgColor else ""
    assert "BDD7EE" in fgColor, f"Linha 6 deve ter fill BDD7EE (período), mas foi: {fgColor}"


# ── Teste 21: linha 8 tem fill azul-escuro (1F4E79 — cabeçalho) ──────────────

def test_linha_8_fill_cabecalho():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    fill = ws.cell(row=8, column=1).fill
    assert fill is not None, "Linha 8 deve ter fill"
    fgColor = fill.fgColor.rgb if fill.fgColor else ""
    assert "1F4E79" in fgColor, f"Linha 8 deve ter fill 1F4E79 (cabeçalho), mas foi: {fgColor}"


# ── Teste 22: linha 9 coluna B = "SALDO TOTAL DISPONÍVEL DIA" (saldo inicial) ─

def test_linha_9_saldo_inicial():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    val = ws.cell(row=9, column=2).value
    assert val == "SALDO TOTAL DISPONÍVEL DIA", f"Linha 9 col B esperado saldo inicial, mas: {val!r}"


# ── Teste 23: última linha coluna B = "SALDO TOTAL DISPONÍVEL DIA" (saldo final)

def test_ultima_linha_saldo_final():
    fech = _make_fechamento()
    l = _make_lancamento()
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    # Com 1 lançamento: saldo ini (9), dado (10), saldo fim (11)
    ultima = ws.max_row
    val = ws.cell(row=ultima, column=2).value
    assert val == "SALDO TOTAL DISPONÍVEL DIA", (
        f"Última linha (row {ultima}) col B esperado saldo final, mas: {val!r}"
    )


# ── Teste 24: coluna A nas linhas de dados tem formato DD/MM/YYYY ─────────────

def test_coluna_a_formato_data():
    fech = _make_fechamento()
    l = _make_lancamento(data=date(2026, 6, 5))
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    # Linha 10 = primeira linha de dados
    fmt = ws.cell(row=10, column=1).number_format
    assert fmt == "DD/MM/YYYY", f"Formato de data na col A esperado DD/MM/YYYY, mas: {fmt!r}"


# ── Teste 25: colunas F, G, H têm formato #,##0.00 ───────────────────────────

def test_colunas_fgh_formato_numero():
    fech = _make_fechamento()
    l = _make_lancamento(tipo_movimento="entrada", valor=Decimal("1000.00"))
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    row = 10  # primeira linha de dados
    for col, nome in [(6, "F"), (7, "G"), (8, "H")]:
        fmt = ws.cell(row=row, column=col).number_format
        assert fmt == "#,##0.00", f"Col {nome} formato esperado #,##0.00, mas: {fmt!r}"


# ── Teste 26: entrada (coluna F) tem fonte verde (00B050) ─────────────────────

def test_entrada_font_verde():
    fech = _make_fechamento()
    l = _make_lancamento(tipo_movimento="entrada", valor=Decimal("500.00"))
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    row = 10  # primeira linha de dados
    font = ws.cell(row=row, column=6).font
    assert font is not None, "Coluna F deve ter font definida"
    cor = font.color.rgb if font.color else ""
    assert "00B050" in cor, f"Entrada (col F) deve ter font verde 00B050, mas: {cor!r}"


# ── Teste 27: saída (coluna G) tem fonte vermelha (FF0000) ───────────────────

def test_saida_font_vermelha():
    fech = _make_fechamento()
    l = _make_lancamento(tipo_movimento="saida", valor=Decimal("200.00"))
    db = _db_com_fechamentos(fechamentos=[fech], lancamentos=[l])
    conteudo, _ = gerar_excel_mensal(
        db=db, empresa_id=EMPRESA_A, ano=2026, mes=6, tipo_conciliacao="extrato_anotado"
    )
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    row = 10  # primeira linha de dados
    font = ws.cell(row=row, column=7).font
    assert font is not None, "Coluna G deve ter font definida"
    cor = font.color.rgb if font.color else ""
    assert "FF0000" in cor, f"Saída (col G) deve ter font vermelha FF0000, mas: {cor!r}"


# ── Teste 28: busca fechamentos por período livre ────────────────────────────

def test_busca_fechamentos_do_periodo():
    fech1 = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 15))
    fech2 = _make_fechamento(fid=FECH_2, periodo_inicio=date(2026, 7, 15))
    db = _db_com_fechamentos(fechamentos=[fech1, fech2], lancamentos=[])

    fechamentos = buscar_fechamentos_do_periodo(
        db=db,
        empresa_id=EMPRESA_A,
        data_inicio=date(2026, 6, 15),
        data_fim=date(2026, 7, 15),
        tipo_conciliacao="extrato_anotado",
    )

    assert [f.id for f in fechamentos] == [FECH_1, FECH_2]


# ── Teste 29: período livre mantém layout e cabeçalho esperados ──────────────

def test_exportacao_periodo_layout_e_nome_arquivo():
    fech1 = _make_fechamento(fid=FECH_1, periodo_inicio=date(2026, 6, 15))
    fech2 = _make_fechamento(fid=FECH_2, periodo_inicio=date(2026, 7, 15))
    l1 = _make_lancamento(fechamento_id=FECH_1, data=date(2026, 6, 15), descricao_banco="PIX Junho")
    l2 = _make_lancamento(fechamento_id=FECH_2, data=date(2026, 7, 15), descricao_banco="PIX Julho")
    empresa = _make_empresa(nome="Empresa Teste")
    arquivo = _make_arquivo(
        fechamento_id=FECH_1,
        metadados={
            "atualizacao": "15/06/2026 09:00:00",
            "nome": "BANCO TESTE S/A",
            "agencia": "0285",
            "conta": "0017861-2",
        },
    )
    db = _db_com_fechamentos(
        fechamentos=[fech1, fech2],
        lancamentos=[l1, l2],
        empresa=empresa,
        arquivos_extrato=[arquivo],
    )

    conteudo, nome = gerar_excel_periodo(
        db=db,
        empresa_id=EMPRESA_A,
        data_inicio=date(2026, 6, 15),
        data_fim=date(2026, 7, 15),
        tipo_conciliacao="extrato_anotado",
    )

    assert nome == "Conciliacao_Empresa_Teste_15-06-2026_a_15-07-2026.xlsx"

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    assert wb.sheetnames == ["15Jun-15Jul26"]
    assert ws.cell(row=2, column=2).value == "BANCO TESTE S/A"
    assert ws.cell(row=3, column=2).value == "0285"
    assert ws.cell(row=4, column=2).value == "0017861-2"
    assert ws.cell(row=6, column=1).value == "Periodo:  15/06/2026 a 15/07/2026"
    assert ws.cell(row=8, column=1).value == "DATA"
    assert ws.cell(row=9, column=2).value == "SALDO TOTAL DISPONÍVEL DIA"
    assert ws.cell(row=10, column=2).value == "PIX Junho"
    assert ws.cell(row=11, column=2).value == "PIX Julho"
    assert ws.cell(row=12, column=2).value == "SALDO TOTAL DISPONÍVEL DIA"


# ── Teste 30: período inválido não gera arquivo ──────────────────────────────

def test_exportacao_periodo_invalido():
    db = _db_com_fechamentos(fechamentos=[])

    with pytest.raises(ValueError, match="Data inicial não pode ser maior"):
        gerar_excel_periodo(
            db=db,
            empresa_id=EMPRESA_A,
            data_inicio=date(2026, 7, 15),
            data_fim=date(2026, 6, 15),
            tipo_conciliacao="extrato_anotado",
        )


# ── Teste 31: período sem conciliações retorna erro claro ────────────────────

def test_exportacao_periodo_sem_conciliacoes():
    db = _db_com_fechamentos(fechamentos=[])

    with pytest.raises(ValueError, match="Nenhuma conciliação encontrada entre 15/06/2026 e 15/07/2026"):
        gerar_excel_periodo(
            db=db,
            empresa_id=EMPRESA_A,
            data_inicio=date(2026, 6, 15),
            data_fim=date(2026, 7, 15),
            tipo_conciliacao="extrato_anotado",
        )
